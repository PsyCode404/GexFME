# Updated: 2025-06-27 11:40:00
import os
import logging
import json
import datetime
import shutil
import tempfile
import ezdxf
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
from flask import request, jsonify, send_file
import io
import math

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def create_folder():
    """Crée un dossier utilisateur basé sur l'email fourni dans la requête"""
    logger.info("Requête POST reçue pour créer un dossier utilisateur")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Récupérer les données de la requête
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe, sinon le créer
        if not os.path.exists(resource_dir):
            logger.info(f"Création du dossier Ressources: {resource_dir}")
            os.makedirs(resource_dir)
        
        # Créer le nom du dossier utilisateur basé sur l'email
        folder_name = email.split('@')[0]
        resource_path = os.path.join(resource_dir, folder_name)
        
        # Vérifier si le dossier existe déjà
        if os.path.exists(resource_path):
            logger.info(f"Le dossier existe déjà: {resource_path}")
            return jsonify({
                'message': 'Le dossier existe déjà',
                'folderName': folder_name,
                'folderExists': True
            }), 200
        
        # Créer le dossier
        os.makedirs(resource_path)
        logger.info(f"Dossier créé avec succès: {resource_path}")
        
        # Ajouter l'entrée dans la base de données en communiquant avec le service principal
        try:
            # Récupérer l'ID utilisateur à partir de l'email
            user_data = {
                'email': email,
                'folder_name': folder_name
            }
            
            # Appeler le service principal pour ajouter l'entrée dans la base de données
            # Comme nous sommes maintenant intégrés dans le même backend, nous pouvons utiliser
            # un chemin relatif ou importer directement la fonction
            from app import app
            with app.app_context():
                from app.controllers.user_routes import register_folder
                result = register_folder(user_data)
                response_status = 200 if result.get('success') else 400
            
            if response_status == 200:
                logger.info(f"Entrée ajoutée dans la base de données pour le dossier: {folder_name}")
            else:
                logger.warning(f"Impossible d'ajouter l'entrée dans la base de données")
        except Exception as e:
            logger.error(f"Erreur lors de l'ajout de l'entrée dans la base de données: {str(e)}")
        
        return jsonify({
            'message': 'Dossier créé avec succès',
            'folderName': folder_name,
            'folderExists': True
        }), 201
    
    except Exception as e:
        logger.error(f"Erreur lors de la création du dossier: {str(e)}")
        return jsonify({'error': f'Erreur lors de la création du dossier: {str(e)}'}), 500

def check_folder():
    """Vérifie si un dossier utilisateur existe basé sur l'email fourni dans la requête"""
    logger.info("Requête POST reçue pour vérifier un dossier utilisateur")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Récupérer les données de la requête
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe
        if not os.path.exists(resource_dir):
            logger.info(f"Le dossier Ressources n'existe pas: {resource_dir}")
            return jsonify({
                'folderExists': False,
                'folderName': email.split('@')[0],
                'message': 'Le dossier Ressources n\'existe pas'
            }), 200
        
        # Vérifier si le dossier utilisateur existe
        folder_name = email.split('@')[0]
        resource_path = os.path.join(resource_dir, folder_name)
        folder_exists = os.path.exists(resource_path)
        
        logger.info(f"Vérification du dossier: {resource_path}, existe: {folder_exists}")
        return jsonify({
            'folderExists': folder_exists,
            'folderName': folder_name,
            'message': 'Dossier vérifié avec succès'
        }), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de la vérification du dossier: {str(e)}")
        return jsonify({'error': f'Erreur lors de la vérification du dossier: {str(e)}'}), 500

def get_folder_structure(folder_path):
    """Récupère la structure des fichiers et dossiers à partir d'un chemin donné"""
    structure = {
        'folders': [],
        'files': []
    }
    
    try:
        if not os.path.exists(folder_path):
            logger.warning(f"Le dossier n'existe pas: {folder_path}")
            return structure
        
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)
            
            if os.path.isdir(item_path):
                # C'est un dossier
                sub_structure = get_folder_structure(item_path)  # Récupérer récursivement la structure du sous-dossier
                folder_info = {
                    'name': item,
                    'path': os.path.relpath(item_path, folder_path),
                    'last_modified': datetime.datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S'),
                    'sub_structure': sub_structure  # Ajouter la sous-structure au dossier
                }
                structure['folders'].append(folder_info)
            else:
                # C'est un fichier
                file_size = os.path.getsize(item_path)
                structure['files'].append({
                    'name': item,
                    'path': os.path.relpath(item_path, folder_path),
                    'size': file_size,
                    'size_formatted': format_file_size(file_size),
                    'last_modified': datetime.datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
        
        # Trier les dossiers et fichiers par nom
        structure['folders'].sort(key=lambda x: x['name'].lower())
        structure['files'].sort(key=lambda x: x['name'].lower())
        
        return structure
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la structure du dossier: {str(e)}")
        return structure

def format_file_size(size_in_bytes):
    """Formate la taille d'un fichier en unités lisibles"""
    if size_in_bytes < 1024:
        return f"{size_in_bytes} B"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.2f} KB"
    elif size_in_bytes < 1024 * 1024 * 1024:
        return f"{size_in_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_in_bytes / (1024 * 1024 * 1024):.2f} GB"

def get_folder_files():
    """Récupère les fichiers du dossier utilisateur"""
    logger.info("Requête POST reçue pour récupérer les fichiers du dossier utilisateur")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Récupérer les données de la requête
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe
        if not os.path.exists(resource_dir):
            logger.info(f"Le dossier Ressources n'existe pas: {resource_dir}")
            return jsonify({
                'folders': [],
                'files': [],
                'message': 'Le dossier Ressources n\'existe pas'
            }), 200
        
        # Vérifier si le dossier utilisateur existe
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        if not os.path.exists(user_folder_path):
            logger.info(f"Le dossier utilisateur n'existe pas: {user_folder_path}")
            return jsonify({
                'folders': [],
                'files': [],
                'message': 'Le dossier utilisateur n\'existe pas'
            }), 200
        
        # Récupérer la structure du dossier
        folder_structure = get_folder_structure(user_folder_path)
        logger.info(f"Structure du dossier récupérée: {json.dumps(folder_structure, indent=2)}")
        
        return jsonify(folder_structure), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des fichiers: {str(e)}")
        return jsonify({'error': f'Erreur lors de la récupération des fichiers: {str(e)}'}), 500

def transfer_files():
    """Transfère les fichiers dans le dossier utilisateur"""
    logger.info("Requête POST reçue pour le transfert de fichiers")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Vérifier si les fichiers sont présents dans la requête
        if 'file1' not in request.files or 'file2' not in request.files:
            logger.error("Deux fichiers sont requis")
            return jsonify({"error": "Deux fichiers sont requis pour le transfert"}), 400

        # Récupérer les fichiers et les noms de fichiers
        file1 = request.files['file1']
        file2 = request.files['file2']
        filename1 = request.form.get('filename1')
        filename2 = request.form.get('filename2')
        custom_folder_name = request.form.get('customFolderName')
        
        # Extraire l'email du formulaire ou de l'en-tête
        email = request.form.get('email')
        
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({"error": "Email non fourni"}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe
        if not os.path.exists(resource_dir):
            os.makedirs(resource_dir)
            logger.info(f"Dossier Ressources créé: {resource_dir}")
        
        # Déterminer le nom du dossier utilisateur
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        # Vérifier si le dossier utilisateur existe
        if not os.path.exists(user_folder_path):
            os.makedirs(user_folder_path)
            logger.info(f"Dossier utilisateur créé: {user_folder_path}")
        
        # Créer le dossier de transfert (sous-dossier dans le dossier utilisateur)
        if custom_folder_name:
            transfer_folder = os.path.join(user_folder_path, secure_filename(custom_folder_name))
        else:
            transfer_folder = os.path.join(user_folder_path, 'Transfert')
        
        # Vérifier si le dossier de transfert existe
        if not os.path.exists(transfer_folder):
            os.makedirs(transfer_folder)
            logger.info(f"Dossier de transfert créé: {transfer_folder}")
        
        # Sécuriser les noms de fichiers
        if not filename1:
            filename1 = secure_filename(file1.filename)
        else:
            filename1 = secure_filename(filename1)
        
        if not filename2:
            filename2 = secure_filename(file2.filename)
        else:
            filename2 = secure_filename(filename2)
        
        # Enregistrer les fichiers
        file1_path = os.path.join(transfer_folder, filename1)
        file2_path = os.path.join(transfer_folder, filename2)
        
        file1.save(file1_path)
        file2.save(file2_path)
        logger.debug(f"Fichiers sauvegardés: {file1_path}, {file2_path}")
        
        return jsonify({"message": f"Fichiers transférés avec succès dans {os.path.basename(transfer_folder)}"}), 200
    
    except Exception as e:
        logger.error(f"Erreur lors du transfert des fichiers: {str(e)}")
        return jsonify({"error": f"Erreur lors du transfert des fichiers: {str(e)}"}), 500

def extract_file_data(file):
    """Extrait les données d'un fichier DXF"""
    temp_file_path = None
    try:
        logger.info(f"Début de l'extraction pour le fichier : {file.filename}")
        
        # Créer un fichier temporaire avec un chemin explicite et mode binaire
        with tempfile.NamedTemporaryFile(delete=False, suffix='.dxf', mode='wb') as temp_file:
            file.save(temp_file)
            temp_file_path = temp_file.name
            logger.info(f"Fichier temporaire sauvegardé : {temp_file_path}")
            
        # Charger le fichier DXF à partir du chemin temporaire
        logger.info(f"Lecture du fichier DXF : {temp_file_path}")
        doc = ezdxf.readfile(temp_file_path)
        
        # Extraire les calques (layers)
        layers = [
            {
                "name": layer.dxf.name,
                "color": layer.dxf.color if layer.dxf.color != 0 else 'N/A',
                "lineweight": layer.dxf.lineweight if hasattr(layer.dxf, 'lineweight') else None
            }
            for layer in doc.layers 
            if not layer.dxf.name.startswith('*')  # Exclure les calques système
        ]
        
        # Extraire les entités (polylignes, lignes, cercles, arcs, texte, etc.)
        modelspace = doc.modelspace()
        
        polylines = []
        lines = []
        circles = []
        arcs = []
        texts = []
        
        for entity in modelspace.query('*'):
            dxftype = entity.dxftype()
            if dxftype == 'POLYLINE':
                polylines.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'vertices': [{'x': v[0], 'y': v[1]} for v in entity.points()],
                    'closed': entity.closed,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'LWPOLYLINE':
                polylines.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'vertices': [{'x': v[0], 'y': v[1]} for v in entity.get_points()],
                    'closed': entity.closed,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'LINE':
                lines.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'start': {'x': entity.dxf.start[0], 'y': entity.dxf.start[1]},
                    'end': {'x': entity.dxf.end[0], 'y': entity.dxf.end[1]},
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'CIRCLE':
                circles.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'center': {'x': entity.dxf.center[0], 'y': entity.dxf.center[1]},
                    'radius': entity.dxf.radius,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'ARC':
                arcs.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'center': {'x': entity.dxf.center[0], 'y': entity.dxf.center[1]},
                    'radius': entity.dxf.radius,
                    'start_angle': entity.dxf.start_angle,
                    'end_angle': entity.dxf.end_angle,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'TEXT':
                texts.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'text': entity.dxf.text,
                    'position': {'x': entity.dxf.insert[0], 'y': entity.dxf.insert[1]},
                    'height': entity.dxf.height,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
        
        # Statistiques globales
        total_entities = len(modelspace)
        statistics = {
            "layer_count": len(layers),
            "polyline_count": len(polylines),
            "line_count": len(lines),
            "circle_count": len(circles),
            "arc_count": len(arcs),
            "text_count": len(texts),
            "total_entities": total_entities
        }
        
        logger.info("Données extraites avec succès")
        return {
            "layers": layers,
            "polylines": polylines,
            "lines": lines,
            "circles": circles,
            "arcs": arcs,
            "texts": texts,
            "statistics": statistics
        }
    
    except Exception as e:
        logger.error(f"Erreur lors de l'extraction des données : {str(e)}")
        return {"error": f"Erreur lors de l'extraction des données : {str(e)}"}
    
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.info(f"Fichier temporaire supprimé : {temp_file_path}")
            except Exception as e:
                logger.error(f"Erreur lors de la suppression du fichier temporaire : {str(e)}")

def extract_data_from_file():
    """Extrait les données d'un fichier DXF dans le dossier de l'utilisateur"""
    logger.info("Requête POST reçue pour extraire les données d'un fichier DXF")
    
    try:
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        filename = data.get('filename')
        folder = data.get('folder', "")
        email = data.get('email')
        file_type = data.get('fileType', 'projet')
        
        logger.info(f"Type de fichier: {file_type}")
        
        if not filename:
            logger.error("Nom de fichier manquant")
            return jsonify({"error": "Nom de fichier requis"}), 400
        
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({"error": "Email non fourni"}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        logger.info(f"Nom du fichier: {filename}")
        logger.info(f"Dossier: {folder}")
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        if not os.path.exists(resource_dir):
            logger.error(f"Le dossier Ressources n'existe pas: {resource_dir}")
            return jsonify({"error": "Dossier Ressources non trouvé"}), 400
        
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        if not os.path.exists(user_folder_path):
            logger.error(f"Le dossier utilisateur n'existe pas: {user_folder_path}")
            return jsonify({"error": "Dossier utilisateur non trouvé"}), 400
        
        logger.info(f"Construction du chemin pour le fichier: {filename}, dossier: {folder}")
        
        if folder:
            file_path = os.path.join(user_folder_path, folder, filename)
        else:
            file_path = os.path.join(user_folder_path, filename)
            
        logger.info(f"Tentative d'accès au fichier: {file_path}")
        
        if os.path.exists(file_path):
            logger.info(f"Fichier trouvé: {file_path}")
        else:
            logger.warning(f"Fichier non trouvé au chemin exact: {file_path}, recherche alternative...")
            
            for root, dirs, files in os.walk(user_folder_path):
                if filename in files:
                    alt_file_path = os.path.join(root, filename)
                    logger.info(f"Fichier trouvé à un emplacement alternatif: {alt_file_path}")
                    file_path = alt_file_path
                    break
            
            if not os.path.exists(file_path):
                logger.error(f"Le fichier n'existe pas après recherche approfondie: {file_path}")
                found_files = []
                for root, dirs, files in os.walk(user_folder_path):
                    for file in files:
                        found_files.append(os.path.join(root, file))
                logger.info(f"Fichiers trouvés dans le dossier utilisateur: {found_files}")
                return jsonify({"error": "Fichier non trouvé"}), 400
        
        logger.info(f"Vérification du fichier: {file_path}")
        logger.info(f"Extension du fichier: {os.path.splitext(file_path)[1]}")
        
        if not file_path.lower().endswith('.dxf'):
            logger.warning(f"Format non standard détecté: {file_path} - tentative d'extraction quand même")
        
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        file_obj = FileStorage(
            stream=io.BytesIO(file_content),
            filename=os.path.basename(file_path),
            content_type='application/octet-stream'
        )
        
        extracted_data = extract_file_data(file_obj)
        logger.info(f"Données extraites avec succès pour {file_path} (Type: {file_type})")
        
        extracted_data['fileType'] = file_type
        
        rel_path = os.path.relpath(file_path, user_folder_path)
        logger.info(f"Chemin relatif du fichier: {rel_path}")
        
        path_components = rel_path.split(os.sep)
        
        if len(path_components) > 1:
            parent_folder = path_components[0]
            logger.info(f"Dossier parent détecté: {parent_folder}")
            extracted_data['sourcePath'] = rel_path
            extracted_data['parentFolder'] = parent_folder
        else:
            extracted_data['sourcePath'] = rel_path
            extracted_data['parentFolder'] = ''
        
        return jsonify(extracted_data), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de l'extraction: {str(e)}")
        return jsonify({"error": f"Erreur lors de l'extraction: {str(e)}"}), 500

def get_visa_content():
    """Récupère le contenu d'un fichier visa.txt"""
    logger.info("Requête POST reçue pour récupérer le contenu d'un fichier visa")
    
    try:
        data = request.get_json() or {}
        file_path = data.get('filePath')
        email = data.get('email')
        
        logger.info(f"Tentative d'accès au fichier: {file_path}")
        
        if email and '/' in file_path and not file_path.startswith('/'):
            current_dir = os.path.dirname(os.path.abspath(__file__))
            resource_dir = os.path.join(current_dir, 'app', 'Ressources')
            folder_name = email.split('@')[0]
            user_folder_path = os.path.join(resource_dir, folder_name)
            
            complete_file_path = os.path.join(user_folder_path, file_path)
            logger.info(f"Chemin complet construit: {complete_file_path}")
            file_path = complete_file_path
        
        if not file_path or not os.path.exists(file_path):
            logger.error(f"Fichier non trouvé: {file_path}")
            
            if email:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                resource_dir = os.path.join(current_dir, 'app', 'Ressources')
                folder_name = email.split('@')[0]
                user_folder_path = os.path.join(resource_dir, folder_name)
                
                if '/' in file_path:
                    filename = file_path.split('/')[-1]
                else:
                    filename = file_path
                    
                logger.info(f"Recherche du fichier {filename} dans le dossier utilisateur {user_folder_path}")
                
                found = False
                for root, dirs, files in os.walk(user_folder_path):
                    if filename in files:
                        file_path = os.path.join(root, filename)
                        found = True
                        logger.info(f"Fichier trouvé à: {file_path}")
                        break
                
                if not found:
                    logger.error(f"Fichier non trouvé après recherche approfondie")
                    return jsonify({'error': 'Fichier non trouvé'}), 404
            else:
                return jsonify({'error': 'Fichier non trouvé'}), 404
        
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        logger.info(f"Contenu récupéré avec succès: {file_path}")
        
        return jsonify({
            'message': 'Contenu récupéré avec succès',
            'content': content
        }), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de la récupération du contenu: {str(e)}")
        return jsonify({'error': f'Erreur lors de la récupération du contenu: {str(e)}'}), 500

def download_visa_file():
    """Permet le téléchargement d'un fichier visa.txt"""
    logger.info("Requête GET reçue pour télécharger un fichier visa")
    
    try:
        file_path = request.args.get('filePath')
        email = request.args.get('email')
        
        logger.info(f"Paramètres reçus - filePath: {file_path}, email: {email}")
        
        if email and file_path:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            resource_dir = os.path.join(current_dir, 'app', 'Ressources')
            folder_name = email.split('@')[0]
            user_folder_path = os.path.join(resource_dir, folder_name)
            
            complete_file_path = os.path.join(user_folder_path, file_path)
            logger.info(f"Chemin complet construit: {complete_file_path}")
            file_path = complete_file_path
        
        if not file_path or not os.path.exists(file_path):
            logger.error(f"Fichier non trouvé: {file_path}")
            
            if email:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                resource_dir = os.path.join(current_dir, 'app', 'Ressources')
                folder_name = email.split('@')[0]
                user_folder_path = os.path.join(resource_dir, folder_name)
                
                if '/' in file_path:
                    filename = file_path.split('/')[-1]
                else:
                    filename = file_path
                    
                logger.info(f"Recherche du fichier {filename} dans le dossier utilisateur {user_folder_path}")
                
                found = False
                for root, dirs, files in os.walk(user_folder_path):
                    if filename in files:
                        file_path = os.path.join(root, filename)
                        found = True
                        logger.info(f"Fichier trouvé à: {file_path}")
                        break
                
                if not found:
                    return jsonify({'error': 'Fichier non trouvé'}), 404
            else:
                return jsonify({'error': 'Fichier non trouvé'}), 404
        
        file_name = os.path.basename(file_path)
        
        logger.info(f"Téléchargement du fichier: {file_path}")
        return send_file(
            file_path,
            mimetype='text/plain',
            as_attachment=True,
            download_name=file_name
        )
    
    except Exception as e:
        logger.error(f"Erreur lors du téléchargement du fichier: {str(e)}")
        return jsonify({'error': f'Erreur lors du téléchargement du fichier: {str(e)}'}), 500

def generate_visa_file():
    """Génère un fichier visa.txt avec les informations de surface calculées"""
    logger.info("Requête POST reçue pour générer un fichier visa.txt")
    
    try:
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        surfaces = data.get('surfaces')
        floor_name = data.get('floorName', 'Sans nom')
        folder_path = data.get('folderPath', '')
        
        logger.info(f"Dossier parent pour le fichier visa: {folder_path}")
        
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        if not surfaces:
            logger.error("Données de surface non fournies")
            return jsonify({'error': 'Données de surface non fournies'}), 400
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        if not os.path.exists(resource_dir):
            logger.info(f"Création du dossier Ressources: {resource_dir}")
            os.makedirs(resource_dir)
        
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        if not os.path.exists(user_folder_path):
            logger.info(f"Création du dossier utilisateur: {user_folder_path}")
            os.makedirs(user_folder_path)
        
        target_folder_path = user_folder_path
        
        if folder_path:
            target_folder_path = os.path.join(user_folder_path, folder_path)
            if not os.path.exists(target_folder_path):
                logger.warning(f"Le dossier {folder_path} n'existe pas, création: {target_folder_path}")
                os.makedirs(target_folder_path)
            else:
                logger.info(f"Utilisation du dossier existant: {target_folder_path}")
        
        output_folder_path = os.path.join(target_folder_path, 'Output')
        if not os.path.exists(output_folder_path):
            logger.info(f"Création du dossier Output: {output_folder_path}")
            os.makedirs(output_folder_path)
        
        file_name = f"visa_{floor_name.replace(' ', '_')}.txt"
        file_path = os.path.join(output_folder_path, file_name)
        
        content = generate_visa_content(surfaces, floor_name)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"Fichier visa généré avec succès: {file_path}")
        
        return jsonify({
            'message': 'Fichier visa généré avec succès',
            'filePath': file_path
        }), 201
    
    except Exception as e:
        logger.error(f"Erreur lors de la génération du fichier visa: {str(e)}")
        return jsonify({'error': f'Erreur lors de la génération du fichier visa: {str(e)}'}), 500

def generate_excel_file():
    """Génère un fichier Excel avec les informations de surface calculées"""
    logger.info("Requête POST reçue pour générer un fichier Excel")
    
    try:
        # Journal détaillé des étapes pour déboguer
        logger.info("Début du traitement de la requête generate-excel-file")
        
        # Détail complet de la requête
        try:
            data = request.get_json() or {}
            logger.info(f"Données reçues (brut): {data}")
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des données JSON: {str(e)}")
            return jsonify({'error': f'Erreur lors de la récupération des données JSON: {str(e)}'}), 500
        
        # Extraction et validation des données
        try:
            email = data.get('email', '')
            surfaces = data.get('surfaces', {})
            floor_name = data.get('floorName', 'Sans nom')
            folder_path = data.get('folderPath', '')
            
            logger.info(f"Données extraites - email: {email}, floor_name: {floor_name}, folder_path: {folder_path}")
            logger.info(f"Structure de 'surfaces': {list(surfaces.keys()) if isinstance(surfaces, dict) else 'NON-DICT'}")
            
            if not email:
                logger.error("Email non fourni dans la requête")
                return jsonify({'error': 'Email non fourni'}), 400
            
            if not surfaces:
                logger.error("Données des surfaces non fournies")
                return jsonify({'error': 'Données des surfaces non fournies'}), 400
        except Exception as e:
            logger.error(f"Erreur lors de la validation des données: {str(e)}")
            return jsonify({'error': f'Erreur lors de la validation des données: {str(e)}'}), 500
        
        # Préparation des dossiers
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            folder_name = email.split('@')[0]
            resource_dir = os.path.join(current_dir, 'app', 'Ressources', folder_name)
            
            logger.info(f"Chemins - current_dir: {current_dir}, folder_name: {folder_name}, resource_dir: {resource_dir}")
            
            if not os.path.exists(resource_dir):
                logger.warning(f"Le dossier de l'utilisateur n'existe pas: {resource_dir}, tentative de création")
                os.makedirs(resource_dir, exist_ok=True)
            
            if folder_path:
                project_dir = os.path.join(resource_dir, folder_path)
                if not os.path.exists(project_dir):
                    logger.info(f"Création du dossier projet: {project_dir}")
                    os.makedirs(project_dir, exist_ok=True)
                
                output_dir = os.path.join(project_dir, 'Output')
            else:
                output_dir = os.path.join(resource_dir, 'Output')
            
            if not os.path.exists(output_dir):
                logger.info(f"Création du dossier Output: {output_dir}")
                os.makedirs(output_dir, exist_ok=True)
                
            logger.info(f"Dossier Output finalisé: {output_dir} (existe: {os.path.exists(output_dir)})")
        except Exception as e:
            logger.error(f"Erreur lors de la préparation des dossiers: {str(e)}")
            return jsonify({'error': f'Erreur lors de la préparation des dossiers: {str(e)}'}), 500
        
        # Génération du fichier Excel
        try:
            # Sanitize du nom d'étage pour le nom de fichier
            sanitized_floor_name = ''.join(c if c.isalnum() or c in '-_' else '_' for c in floor_name)
            now = datetime.datetime.now()
            date_str = now.strftime('%Y%m%d_%H%M%S')
            excel_filename = f"surface_comparison_{sanitized_floor_name}_{date_str}.xlsx"
            excel_path = os.path.join(output_dir, excel_filename)
            
            logger.info(f"Préparation de la création du fichier Excel: {excel_path}")
            
            # Remplacer la fonction create_excel_document par une version très simplifiée pour tester
            # Cette version crée un fichier Excel minimal fonctionnel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SDP"
            
            # En-têtes complets selon l'image de référence
            ws['A1'] = "Etages"
            ws['B1'] = "Destinations"
            ws['C1'] = "Surface existante avant travaux (A)"
            ws['D1'] = "Surface creee (B)"
            ws['E1'] = "Surface creee par changement de destination"
            ws['F1'] = "Surface demolie reconstruite"
            ws['G1'] = "Surface supprimee (D)"
            ws['H1'] = "Surface supprimee par changement de destination"
            ws['I1'] = "Surface projet"
            ws['J1'] = "Surface RDV"
            
            # Ajout de l'étage - seulement dans la première ligne
            ws['A2'] = floor_name
            
            # Traitement très simplifié des destinations
            row = 2
            
            # Fonction pour calculer l'intersection entre deux polylignes
            def calculate_intersection(poly1, poly2):
                try:
                    # Convertir les sommets en objets Point
                    from shapely.geometry import Polygon
                    
                    # Créer les polygones à partir des sommets
                    poly1_pts = [(v['x'], v['y']) for v in poly1.get('vertices', [])]
                    poly2_pts = [(v['x'], v['y']) for v in poly2.get('vertices', [])]
                    
                    if len(poly1_pts) < 3 or len(poly2_pts) < 3:
                        return 0.0
                        
                    poly1_shapely = Polygon(poly1_pts)
                    poly2_shapely = Polygon(poly2_pts)
                    
                    # Vérifier si les polygones sont valides
                    if not poly1_shapely.is_valid or not poly2_shapely.is_valid:
                        return 0.0
                    
                    # Calculer l'intersection
                    intersection = poly1_shapely.intersection(poly2_shapely)
                    
                    # Retourner l'aire de l'intersection
                    return intersection.area
                    
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul d'intersection: {str(e)}")
                    return 0.0
            
            # Dictionnaire de correspondance pour les cas spéciaux
            special_cases = {
                "AUTRE_BUREAU": "Autre bureau",
                "AUTRE_CONGRE": "Autre congrès exposition",
                "AUTRE_ENTREP": "Autre entrepôt",
                "AUTRE_INDUST": "Autre industrie",
                "COMMERCE_ART": "Commerce artisanat",
                "COMMERCE_AUT": "Commerce autre hébergement touristique",
                "COMMERCE_CIN": "Commerce cinéma",
                "COMMERCE_DE_": "Commerce de gros",
                "COMMERCE_HOT": "Commerce hôtel",
                "COMMERCE_RES": "Commerce restauration",
                "COMMERCE_SER": "Commerce service accueil clientèle",
                "EXPLOITATIO": "Exploitation forestière",
                "EXPLOITATION": "Exploitation agricole",
                "HABITATION_H": "Habitation hébergement",
                "HABITATION_L": "Habitation logement",
                "SPIC_ADMINIS": "Spic administration",
                "SPIC_ART_SPE": "Spic art spectacle",
                "SPIC_AUTRE": "Spic autre",
                "SPIC_ENSEIGN": "Spic enseignement santé",
                "SPIC_LT": "Spic lt",
                "SPIC_SPORT": "Spic sport"
            }
            
            # Journaliser la structure complète des données pour déboguer
            logger.info(f"Structure détaillée des surfaces: {json.dumps(surfaces, default=str)}")
            
            # Analyse détaillée de la structure des données
            existant_polylines = surfaces.get('existant', {}).get('polylines', [])
            projet_polylines = surfaces.get('projet', {}).get('polylines', [])
            
            logger.info(f"Existant polylines: {len(existant_polylines)}")
            logger.info(f"Projet polylines: {len(projet_polylines)}")
            
            # IMPORTANT: Rappel de l'inversion des fichiers
            # existant_polylines contient les données du fichier "Projet_demoli_feuille_TA.dxf" (surface existante avant travaux)
            # projet_polylines contient les données du fichier "Existant_exmple_demoli.dxf" (surface projet)
            
            # Fonction pour extraire le nom de destination d'une polyligne avec logging
            def get_destination_from_layer(layer):
                if not isinstance(layer, str):
                    logger.warning(f"Layer n'est pas une chaîne: {layer}")
                    return None
                    
                # Vérifier si c'est un calque SDP_1
                if 'GEX_EDS_SDP_1' not in layer:
                    return None
                
                # Journaliser le calque pour débogage
                logger.info(f"Traitement du calque: {layer}")
                
                # Extraire la partie après GEX_EDS_SDP_1-
                try:
                    parts = layer.split('GEX_EDS_SDP_1-')
                    if len(parts) >= 2:
                        raw_destination = parts[1]
                        
                        # Journaliser la destination brute extraite
                        logger.info(f"Destination brute extraite: {raw_destination}")
                        
                        # Nettoyer le nom de la destination de tout caractère indésirable à la fin
                        if raw_destination.endswith('_'):
                            raw_destination = raw_destination[:-1]
                        
                        # Cas spécial pour EXPLOITATIO0 qui apparaît dans votre fichier
                        if "EXPLOITATIO0" in raw_destination:
                            raw_destination = "EXPLOITATIO"
                        
                        # Tentative de correspondance avec les clés connues
                        for key in special_cases.keys():
                            if key == raw_destination:
                                logger.info(f"Match exact: {key}")
                                return key
                        
                        # Si pas de correspondance exacte, faire une recherche partielle
                        for key in special_cases.keys():
                            if key in raw_destination or raw_destination in key:
                                logger.info(f"Match partiel: {key} dans {raw_destination}")
                                return key
                        
                        logger.info(f"Destination finale: {raw_destination}")
                        return raw_destination
                except Exception as e:
                    logger.error(f"Erreur lors de l'extraction de la destination: {str(e)}")
                
                return None
            
            # Fonction pour déterminer si une polyligne est un calque spécial à déduire
            def is_special_layer(layer):
                if not isinstance(layer, str):
                    return False
                    
                patterns = ['GEX_EDS_SDP_2', 'GEX_EDS_SDP_3', 'GEX_EDS_SDP_4', 'GEX_EDS_SDP_5', 'GEX_EDS_SDP_7']
                return any(pattern in layer for pattern in patterns)
            
            # Fonction pour calculer la surface d'une polyligne
            def calculate_area(polyline):
                vertices = polyline.get('vertices', [])
                if len(vertices) < 3:
                    return 0
                    
                # Calculer l'aire en utilisant la formule de Shoelace
                area = 0.0
                n = len(vertices)
                for i in range(n):
                    j = (i + 1) % n
                    area += vertices[i]['x'] * vertices[j]['y']
                    area -= vertices[j]['x'] * vertices[i]['y']
                    
                return abs(area) / 2.0
            
            def is_contained(polyline1, polyline2):
                try:
                    # Utiliser Shapely pour une vérification de contenance géométrique précise
                    from shapely.geometry import Polygon
                    
                    # Créer les polygones à partir des sommets
                    poly1_pts = [(v['x'], v['y']) for v in polyline1.get('vertices', [])]
                    poly2_pts = [(v['x'], v['y']) for v in polyline2.get('vertices', [])]
                    
                    if len(poly1_pts) < 3 or len(poly2_pts) < 3:
                        return False
                        
                    poly1_shapely = Polygon(poly1_pts)
                    poly2_shapely = Polygon(poly2_pts)
                    
                    # Vérifier si les polygones sont valides
                    if not poly1_shapely.is_valid or not poly2_shapely.is_valid:
                        return False
                    
                    # Vérifier si poly1 est entièrement contenu dans poly2
                    # Un polygone est contenu dans un autre si leur intersection est égale au premier polygone
                    return poly2_shapely.contains(poly1_shapely)
                except Exception as e:
                    logger.warning(f"Erreur lors de la vérification de contenance: {str(e)}")
                    return False
                    
            # Structure pour stocker les résultats
            calculation_results = {
                'existant': {},       # Pour les surfaces existantes
                'projet': {},          # Pour les surfaces projet
                'cree_changement': {}, # Surface créée par changement de destination
                'demolie': {},         # Surface démolie reconstruite
                'supprimee': {},       # Surface supprimée (D)
                'supprimee_changement': {}, # Surface supprimée par changement de destination
                'demolition': {}       # Pour les surfaces de démolition par destination
            }
            
            # Identifier les polylignes de démolition (GEX_EDS_TA_SDP_CAHIER_DEMO)
            demolition_polylines = [p for p in existant_polylines 
                                 if 'GEX_EDS_TA_SDP_CAHIER_DEMO' in p.get('layer', '')]
            
            logger.info(f"Nombre de polylignes de démolition trouvées: {len(demolition_polylines)}")
            
            # Traiter les polylignes existantes (Projet_demoli_feuille_TA.dxf)
            main_existant_polylines = []
            special_existant_polylines = []
            
            for polyline in existant_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' in layer:
                    destination = get_destination_from_layer(layer)
                    if destination:
                        main_existant_polylines.append(polyline)
                        area = calculate_area(polyline)
                        if destination not in calculation_results['existant']:
                            calculation_results['existant'][destination] = 0
                            calculation_results['demolition'][destination] = 0.0
                        calculation_results['existant'][destination] += area
                        logger.info(f"Existant: Destination {destination} - ajout surface {area}")
                        
                        # Calculer l'intersection avec les zones de démolition
                        for demo_poly in demolition_polylines:
                            intersection_area = calculate_intersection(polyline, demo_poly)
                            if intersection_area > 0:
                                calculation_results['demolition'][destination] += intersection_area
                                logger.info(f"  - Intersection avec zone de démolition: {intersection_area:.2f} m²")
                                
                elif is_special_layer(layer):
                    special_existant_polylines.append(polyline)
            
            # Traiter les polylignes projet (Existant_exmple_demoli.dxf)
            main_projet_polylines = []
            special_projet_polylines = []
            
            for polyline in projet_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' in layer:
                    destination = get_destination_from_layer(layer)
                    if destination:
                        main_projet_polylines.append(polyline)
                        area = calculate_area(polyline)
                        
                        # Enregistrer les dimensions brutes pour analyse
                        vertices = polyline.get('vertices', [])
                        if vertices:
                            min_x = min(v['x'] for v in vertices)
                            max_x = max(v['x'] for v in vertices)
                            min_y = min(v['y'] for v in vertices)
                            max_y = max(v['y'] for v in vertices)
                            logger.info(f"Dimensions brutes pour {destination}: X({min_x},{max_x}), Y({min_y},{max_y}), Surface={area}")
                            
                        # Vérifier si le calcul semble correct (cas de figure spécifique)
                        # Les erreurs de calcul peuvent venir de points trop proches ou mal définis
                        
                        if destination not in calculation_results['projet']:
                            calculation_results['projet'][destination] = 0
                        calculation_results['projet'][destination] += area
                        logger.info(f"Projet: Destination {destination} - ajout surface {area}")
                elif is_special_layer(layer):
                    special_projet_polylines.append(polyline)
            
            # Déduire les surfaces spéciales des existantes
            for special_polyline in special_existant_polylines:
                area = calculate_area(special_polyline)
                if area <= 0:
                    continue
                    
                for main_polyline in main_existant_polylines:
                    if is_contained(special_polyline, main_polyline):
                        destination = get_destination_from_layer(main_polyline.get('layer', ''))
                        if destination and destination in calculation_results['existant']:
                            calculation_results['existant'][destination] -= area
                            logger.info(f"Existant: Déduction de {area} pour {destination}")
                        break
            
            # Déduire les surfaces spéciales des projets
            for special_polyline in special_projet_polylines:
                area = calculate_area(special_polyline)
                if area <= 0:
                    continue
                    
                for main_polyline in main_projet_polylines:
                    if is_contained(special_polyline, main_polyline):
                        destination = get_destination_from_layer(main_polyline.get('layer', ''))
                        if destination and destination in calculation_results['projet']:
                            calculation_results['projet'][destination] -= area
                            logger.info(f"Projet: Déduction de {area} pour {destination}")
                        break
            
            # IMPORTANT: Recalculer les surfaces projet avec la même méthode que pour les surfaces existantes
            # Nous allons vider et recalculer les surfaces projet
            logger.info("Recalcul des surfaces projet avec la même méthode que pour les surfaces existantes")
            
            # Vider les résultats précédents pour les surfaces projet
            calculation_results['projet'] = {}
            
            # Appliquer exactement la même méthode pour les surfaces projet que pour les surfaces existantes
            # Pour les polylignes principales
            main_projet_polylines = []
            for polyline in projet_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' in layer:
                    destination = get_destination_from_layer(layer)
                    if destination:
                        main_projet_polylines.append(polyline)
                        area = calculate_area(polyline)
                        if destination not in calculation_results['projet']:
                            calculation_results['projet'][destination] = 0
                        calculation_results['projet'][destination] += area
                        logger.info(f"Recalcul Projet: Destination {destination} - ajout surface {area}")
            
            # Pour les polylignes spéciales
            special_projet_polylines = []
            for polyline in projet_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' not in layer and is_special_layer(layer):
                    special_projet_polylines.append(polyline)
            
            # Déduire les surfaces spéciales des surfaces projet
            for special_polyline in special_projet_polylines:
                area = calculate_area(special_polyline)
                if area <= 0:
                    continue
                    
                # Vérifier la contenance dans chaque polyligne principale
                for main_polyline in main_projet_polylines:
                    if is_contained(special_polyline, main_polyline):
                        destination = get_destination_from_layer(main_polyline.get('layer', ''))
                        if destination and destination in calculation_results['projet']:
                            calculation_results['projet'][destination] -= area
                            logger.info(f"Recalcul Projet: Déduction de {area} pour {destination}")
                        break
            
            # Débogage des résultats de calcul finaux
            logger.info(f"===== Résultats de calcul finaux =====")
            for dest, value in calculation_results['existant'].items():
                logger.info(f"Existant - {dest}: {value}")
            for dest, value in calculation_results['projet'].items():
                logger.info(f"Projet - {dest}: {value}")
            
            # Collecter uniquement les destinations qui ont des valeurs dans les calculs
            all_destinations = set()
            
            # Nous n'ajoutons que les destinations qui ont des polylignes dans les fichiers extraits
            for destination in calculation_results['existant'].keys():
                all_destinations.add(destination)
                
            for destination in calculation_results['projet'].keys():
                all_destinations.add(destination)
                
            logger.info(f"Destinations présentes dans les fichiers extraits: {all_destinations}")
            
            # Vérifier si nous avons au moins une destination
            if not all_destinations:
                logger.warning("Aucune destination trouvée dans les fichiers extraits. Vérifiez le contenu des fichiers DXF.")
            
            # Traiter chaque destination
            for destination in sorted(all_destinations):
                formatted_destination = special_cases.get(destination, destination)
                
                # Surface existante (A) - du fichier Projet_demoli_feuille_TA.dxf
                existant_surface = calculation_results['existant'].get(destination, 0)
                
                # Surface projet - du fichier Existant_exmple_demoli.dxf - CONSERVER LA VALEUR BRUTE
                # Ne pas arrondir à ce stade pour avoir le même traitement que les surfaces existantes
                projet_surface = calculation_results['projet'].get(destination, 0)
                
                # Ajouter la ligne au fichier Excel avec précision complète pour les surfaces
                ws[f'B{row}'] = formatted_destination
                
                # Afficher les valeurs exactes dans les logs avant affichage
                logger.info(f"Destination {formatted_destination} - Surface existante: {existant_surface}, Surface projet: {projet_surface}")
                
                # Surface existante avec précision identique au fichier attendu
                if existant_surface > 0:
                    ws[f'C{row}'] = existant_surface
                
                # Surface projet avec précision identique au fichier attendu
                if projet_surface > 0:
                    ws[f'I{row}'] = projet_surface
                
                # Remplir la colonne Surface démolie (F) uniquement si > 0
                demolition_surface = calculation_results['demolition'].get(destination, 0)
                if demolition_surface > 0:
                    ws[f'F{row}'] = demolition_surface
                
                # Journaliser les résultats pour débogage
                logger.info(f"Destination {formatted_destination} - Surface existante: {existant_surface}, "
                           f"Surface projet: {projet_surface}, Surface démolie: {demolition_surface}")
                
                row += 1
            
            # Supprimer la feuille SDP si elle existe déjà
            if "SDP" in wb.sheetnames:
                del wb["SDP"]
                
            # Créer une nouvelle feuille SDP avec la structure demandée
            ws_sdp = wb.create_sheet(title="SDP")
            
            # En-têtes pour la feuille SDP selon les spécifications
            ws_sdp['A1'] = "Étages"
            ws_sdp['B1'] = "Destinations"
            ws_sdp['C1'] = "Surface existante avant travaux (A)"
            ws_sdp['D1'] = "Surface créée (B)"
            ws_sdp['E1'] = "Surface démolie reconstruite"
            ws_sdp['F1'] = "Surface supprimée (D)"
            ws_sdp['G1'] = "Surface supprimée par changement de destination"
            ws_sdp['H1'] = "Surface projet"
            ws_sdp['I1'] = "Surface RDV"
            
            # Appliquer le style aux en-têtes
            header_font = Font(name='Arial', size=11, bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                cell = ws_sdp[f'{col}1']
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws_sdp.column_dimensions[col].width = 22  # Ajuster la largeur des colonnes
            
            # Créer une nouvelle feuille nommée "TA Projet" (anciennement "TA Demoli Detailled")
            ws_ta = wb.create_sheet(title="TA Projet")
            
            # En-têtes pour la feuille TA Projet
            ws_ta['A1'] = "Étages"
            ws_ta['B1'] = "Destinations"
            ws_ta['C1'] = "TA avant déduction"
            ws_ta['D1'] = "Vides"
            ws_ta['E1'] = "Surfaces dont h < 1.80m"
            ws_ta['F1'] = "TA après déduction"
            
            # Appliquer le style aux en-têtes
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws_ta[f'{col}1']
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws_ta.column_dimensions[col].width = 15  # Ajuster la largeur des colonnes
            
            # Ajouter l'étage dans la feuille TA
            ta_row = 2
            ws_ta[f'A{ta_row}'] = floor_name
            
            # Appliquer le style à la cellule de l'étage
            data_font = Font(name='Arial', size=12)
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            cell = ws_ta[f'A{ta_row}']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Fonction pour déterminer si une polyligne est un calque principal SDP_1
            def is_main_sdp_polyline(polyline):
                layer = polyline.get('layer', '')
                return isinstance(layer, str) and layer.startswith('GEX_EDS_SDP_1')
            
            # Fonction pour déterminer si une polyligne est un calque spécial à déduire (SDP_2, SDP_3, etc.)
            def is_special_polyline(polyline):
                layer = polyline.get('layer', '')
                if not isinstance(layer, str):
                    return False
                # Vérifier si c'est un calque spécial à déduire (pas SDP_1)
                patterns = ['GEX_EDS_SDP_2', 'GEX_EDS_SDP_3', 'GEX_EDS_SDP_4', 'GEX_EDS_SDP_5', 'GEX_EDS_SDP_7']
                return any(pattern in layer for pattern in patterns)
            
            # Fonction pour calculer l'intersection entre deux polylignes et retourner la surface d'intersection
            def calculate_intersection_area(polyline1, polyline2):
                try:
                    from shapely.geometry import Polygon, MultiPolygon, mapping
                    from shapely.ops import unary_union
                    import numpy as np
                    
                    # Extraire les informations des polylignes pour le débogage
                    layer1 = polyline1.get('layer', 'inconnu')
                    layer2 = polyline2.get('layer', 'inconnu')
                    logger.info(f"Calcul d'intersection entre: {layer1} et {layer2}")
                    
                    vertices1 = polyline1.get('vertices', [])
                    vertices2 = polyline2.get('vertices', [])
                    
                    if len(vertices1) < 3 or len(vertices2) < 3:
                        logger.warning(f"Pas assez de sommets pour former un polygone: {len(vertices1)} et {len(vertices2)}")
                        return 0.0
                    
                    # Journaliser les sommets pour débogage
                    logger.info(f"Polygone 1 ({layer1}): {len(vertices1)} sommets")
                    logger.info(f"Polygone 2 ({layer2}): {len(vertices2)} sommets")
                    
                    # Convertir les sommets en points pour Shapely avec précision
                    poly1_pts = [(float(v['x']), float(v['y'])) for v in vertices1]
                    poly2_pts = [(float(v['x']), float(v['y'])) for v in vertices2]
                    
                    # S'assurer que les polygones sont fermés (premier et dernier point identiques)
                    if poly1_pts[0] != poly1_pts[-1]:
                        poly1_pts.append(poly1_pts[0])
                    if poly2_pts[0] != poly2_pts[-1]:
                        poly2_pts.append(poly2_pts[0])
                    
                    # Créer les polygones Shapely
                    try:
                        # Créer les polygones avec des coordonnées précises
                        poly1 = Polygon(poly1_pts)
                        poly2 = Polygon(poly2_pts)
                        
                        # Journaliser les aires des polygones
                        logger.info(f"Aire du polygone 1 ({layer1}): {poly1.area}")
                        logger.info(f"Aire du polygone 2 ({layer2}): {poly2.area}")
                        
                        # Vérifier si les polygones sont valides
                        if not poly1.is_valid:
                            logger.warning(f"Polygone 1 ({layer1}) invalide, tentative de réparation")
                            poly1 = poly1.buffer(0)  # Tenter de réparer le polygone
                        if not poly2.is_valid:
                            logger.warning(f"Polygone 2 ({layer2}) invalide, tentative de réparation")
                            poly2 = poly2.buffer(0)  # Tenter de réparer le polygone
                        
                        if not poly1.is_valid:
                            logger.warning(f"Polygone 1 ({layer1}) toujours invalide après réparation")
                            return 0.0
                        if not poly2.is_valid:
                            logger.warning(f"Polygone 2 ({layer2}) toujours invalide après réparation")
                            return 0.0
                        
                        # Ajouter une petite tolérance pour les intersections presque tangentes
                        poly1_buffered = poly1.buffer(0.001)
                        
                        # Calculer l'intersection avec la tolérance
                        if poly1_buffered.intersects(poly2):
                            # Essayer d'abord avec la tolérance
                            intersection = poly1_buffered.intersection(poly2)
                            intersection_area = intersection.area
                            
                            # Si l'intersection est très petite, essayer sans tolérance
                            if intersection_area < 0.01 and poly1.intersects(poly2):
                                intersection = poly1.intersection(poly2)
                                intersection_area = intersection.area
                            
                            logger.info(f"Intersection trouvée entre {layer1} et {layer2}! Aire: {intersection_area}")
                            
                            # Journaliser les coordonnées de l'intersection pour débogage
                            try:
                                intersection_coords = mapping(intersection)
                                logger.info(f"Forme de l'intersection: {intersection.geom_type}")
                            except Exception as e:
                                logger.warning(f"Impossible d'extraire les coordonnées de l'intersection: {str(e)}")
                            
                            return intersection_area
                        else:
                            # Vérifier si les polygones sont très proches
                            distance = poly1.distance(poly2)
                            logger.info(f"Pas d'intersection entre {layer1} et {layer2}. Distance: {distance}")
                            return 0.0
                    except Exception as e:
                        logger.warning(f"Erreur lors de la création des polygones: {str(e)}")
                        return 0.0
                        
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul d'intersection: {str(e)}")
                    return 0.0
            
            # Filtrer les polylignes principales et spéciales
            main_projet_polylines = [p for p in projet_polylines if is_main_sdp_polyline(p)]
            special_projet_polylines = [p for p in projet_polylines if is_special_polyline(p)]
            
            # Collecter toutes les destinations des polylignes principales
            all_destinations = set()
            for polyline in main_projet_polylines:
                destination = get_destination_from_layer(polyline.get('layer', ''))
                if destination:
                    all_destinations.add(destination)
            
            # Dictionnaire pour stocker les données T.A. par destination
            ta_data = {}
            
            # Calculer les données T.A. pour chaque destination
            for destination in all_destinations:
                # Initialiser les valeurs pour cette destination
                planchers_avant_deductions = 0.0  # Planchers avant déductions
                vides = 0.0                       # Vides (stairwells, elevator shafts, etc.)
                surfaces_h_moins_180 = 0.0        # Surfaces dont h < 1.80m
                
                # Calculer Planchers avant déductions - utiliser les surfaces principales
                for polyline in main_projet_polylines:
                    if get_destination_from_layer(polyline.get('layer', '')) == destination:
                        # Calculer la surface de cette polyligne
                        area = 0.0
                        vertices = polyline.get('vertices', [])
                        if len(vertices) >= 3:
                            # Calculer la surface en utilisant la formule de Shoelace (Gauss's area formula)
                            n = len(vertices)
                            for i in range(n):
                                j = (i + 1) % n
                                area += vertices[i]['x'] * vertices[j]['y']
                                area -= vertices[j]['x'] * vertices[i]['y']
                            area = abs(area) / 2.0
                        planchers_avant_deductions += area
                
                # Identifier les vides (GEX_EDS_SDP_2 - TREMIE)
                for polyline in special_projet_polylines:
                    special_layer = polyline.get('layer', '')
                    if not isinstance(special_layer, str):
                        continue
                    
                    if 'GEX_EDS_SDP_2' in special_layer:  # Vides/TREMIE
                        # Extraire les informations du vide pour le débogage
                        void_id = polyline.get('id', 'inconnu')
                        logger.info(f"Traitement du vide {void_id} avec calque {special_layer}")
                        
                        # Calculer d'abord la surface totale du vide
                        void_area = 0.0
                        vertices = polyline.get('vertices', [])
                        
                        if len(vertices) < 3:
                            logger.warning(f"Pas assez de sommets pour le vide {void_id}: {len(vertices)}")
                            continue
                            
                        from shapely.geometry import Polygon, mapping
                        try:
                            # Créer un polygone Shapely pour le vide
                            void_pts = [(float(v['x']), float(v['y'])) for v in vertices]
                            
                            # S'assurer que le polygone est fermé
                            if void_pts[0] != void_pts[-1]:
                                void_pts.append(void_pts[0])
                            
                            # Journaliser les points du vide pour débogage
                            logger.info(f"Vide {void_id}: {len(void_pts)} points")
                            
                            void_polygon = Polygon(void_pts)
                            if not void_polygon.is_valid:
                                logger.warning(f"Polygone de vide {void_id} invalide, tentative de réparation")
                                void_polygon = void_polygon.buffer(0)  # Réparer le polygone si nécessaire
                            
                            if void_polygon.is_valid:
                                void_area = void_polygon.area
                                logger.info(f"Vide {void_id} valide trouvé avec surface: {void_area}")
                                
                                # Journaliser la forme du vide
                                try:
                                    void_coords = mapping(void_polygon)
                                    logger.info(f"Forme du vide {void_id}: {void_polygon.geom_type}")
                                except Exception as e:
                                    logger.warning(f"Impossible d'extraire les coordonnées du vide {void_id}: {str(e)}")
                            else:
                                logger.warning(f"Polygone de vide {void_id} invalide après tentative de réparation")
                                continue
                        except Exception as e:
                            logger.warning(f"Erreur lors de la création du polygone vide {void_id}: {str(e)}")
                            continue
                        
                        # Calculer l'intersection avec les polylignes de cette destination
                        intersection_found = False
                        destination_polylines = [p for p in main_projet_polylines if get_destination_from_layer(p.get('layer', '')) == destination]
                        
                        logger.info(f"Vérification des intersections pour vide {void_id} avec {destination} ({len(destination_polylines)} polylignes)")
                        
                        # Essayer d'abord avec un buffer légèrement plus grand pour capturer les intersections proches
                        void_polygon_buffered = void_polygon.buffer(0.05)  # Augmenter le buffer à 0.05 pour mieux capturer les intersections
                        
                        for main_polyline in destination_polylines:
                            main_id = main_polyline.get('id', 'inconnu')
                            main_layer = main_polyline.get('layer', 'inconnu')
                            logger.info(f"Tentative d'intersection entre vide {void_id} et polyligne {main_id} (calque {main_layer})")
                            
                            # Calculer l'aire d'intersection avec méthode améliorée
                            intersection_area = calculate_intersection_area(polyline, main_polyline)
                            
                            if intersection_area > 0:
                                vides += intersection_area
                                intersection_found = True
                                logger.info(f"Intersection trouvée entre vide {void_id} et {destination} (polyligne {main_id}): {intersection_area}")
                            else:
                                # Essayer avec le buffer si l'intersection directe échoue
                                try:
                                    main_vertices = main_polyline.get('vertices', [])
                                    if len(main_vertices) >= 3:
                                        main_pts = [(float(v['x']), float(v['y'])) for v in main_vertices]
                                        if main_pts[0] != main_pts[-1]:
                                            main_pts.append(main_pts[0])
                                            
                                        main_polygon = Polygon(main_pts)
                                        if main_polygon.is_valid:
                                            # Vérifier l'intersection avec le buffer
                                            if void_polygon_buffered.intersects(main_polygon):
                                                buffer_intersection = void_polygon_buffered.intersection(main_polygon)
                                                buffer_area = buffer_intersection.area
                                                logger.info(f"Intersection avec buffer trouvée: {buffer_area}")
                                                
                                                # Utiliser l'aire du vide originale si l'intersection est significative
                                                if buffer_area > 0.001:  # Seuil abaissé pour capturer plus d'intersections
                                                    # Si l'intersection directe est nulle, utiliser l'aire du vide
                                                    # dans la zone d'intersection du buffer
                                                    intersection_ratio = buffer_area / void_polygon_buffered.area
                                                    estimated_area = void_area * intersection_ratio
                                                    
                                                    # Vérifier si l'estimation est raisonnable
                                                    if estimated_area > 0.001 and estimated_area <= void_area:
                                                        vides += estimated_area
                                                        intersection_found = True
                                                        logger.info(f"Intersection estimée entre vide {void_id} et {destination} (polyligne {main_id}): {estimated_area}")
                                                    else:
                                                        # Essayer une dernière fois avec un calcul direct
                                                        actual_intersection = void_polygon.intersection(main_polygon)
                                                        actual_area = actual_intersection.area
                                                        if actual_area > 0:
                                                            vides += actual_area
                                                            intersection_found = True
                                                            logger.info(f"Intersection directe après buffer entre vide {void_id} et {destination} (polyligne {main_id}): {actual_area}")
                                                        elif buffer_area > 0.1:  # Si l'intersection avec buffer est significative
                                                            # Utiliser une valeur minimale basée sur le buffer
                                                            min_area = min(void_area, buffer_area * 0.5)
                                                            vides += min_area
                                                            intersection_found = True
                                                            logger.info(f"Intersection minimale entre vide {void_id} et {destination} (polyligne {main_id}): {min_area}")
                                                            logger.info(f"  - Aire du vide: {void_area}")
                                                            logger.info(f"  - Aire de l'intersection avec buffer: {buffer_area}")
                                                            logger.info(f"  - Aire minimale utilisée: {min_area}")
                                                            logger.info(f"  - Coordonnées du vide: {void_polygon.wkt}")
                                                            logger.info(f"  - Coordonnées de la destination: {main_polygon.wkt}")
                                                            logger.info(f"  - Coordonnées de l'intersection: {buffer_intersection.wkt}")
                                                            logger.info(f"  - Distance entre polygones: {void_polygon.distance(main_polygon)}")
                                                            logger.info(f"  - Centroide vide: {void_polygon.centroid.wkt}")
                                                            logger.info(f"  - Centroide destination: {main_polygon.centroid.wkt}")
                                                            logger.info(f"  - Distance entre centroides: {void_polygon.centroid.distance(main_polygon.centroid)}")
                                                            
                                                            # Journaliser la valeur finale des vides pour cette destination
                                                            logger.info(f"VALEUR FINALE DES VIDES POUR {destination}: {vides}")
                                                            
                                                            # Forcer la mise à jour dans ta_data pour cette destination
                                                            if destination in ta_data:
                                                                ta_data[destination]['vides'] = vides
                                except Exception as e:
                                    logger.warning(f"Erreur lors de la vérification d'intersection avec buffer: {str(e)}")
                        
                        # Journaliser si aucune intersection n'a été trouvée pour ce vide avec cette destination
                        if not intersection_found:
                            logger.warning(f"Aucune intersection trouvée entre le vide {void_id} ({void_area}) et la destination {destination}")
                
                # Identifier les surfaces dont h < 1.80m (GEX_EDS_SDP_3 - H-180)
                for polyline in special_projet_polylines:
                    special_layer = polyline.get('layer', '')
                    if not isinstance(special_layer, str):
                        continue
                    
                    if 'GEX_EDS_SDP_3' in special_layer:  # H-180
                        # Extraire les informations de la zone h<1.80m pour le débogage
                        h180_id = polyline.get('id', 'inconnu')
                        logger.info(f"Traitement de la zone h<1.80m {h180_id} avec calque {special_layer}")
                        
                        # Calculer d'abord la surface totale de la zone h<1.80m
                        h180_area = 0.0
                        vertices = polyline.get('vertices', [])
                        
                        if len(vertices) < 3:
                            logger.warning(f"Pas assez de sommets pour la zone h<1.80m {h180_id}: {len(vertices)}")
                            continue
                            
                        from shapely.geometry import Polygon, mapping
                        try:
                            # Créer un polygone Shapely pour la zone h<1.80m
                            h180_pts = [(float(v['x']), float(v['y'])) for v in vertices]
                            
                            # S'assurer que le polygone est fermé
                            if h180_pts[0] != h180_pts[-1]:
                                h180_pts.append(h180_pts[0])
                            
                            # Journaliser les points de la zone h<1.80m pour débogage
                            logger.info(f"Zone h<1.80m {h180_id}: {len(h180_pts)} points")
                            
                            h180_polygon = Polygon(h180_pts)
                            if not h180_polygon.is_valid:
                                logger.warning(f"Polygone de zone h<1.80m {h180_id} invalide, tentative de réparation")
                                h180_polygon = h180_polygon.buffer(0)  # Réparer le polygone si nécessaire
                            
                            if h180_polygon.is_valid:
                                h180_area = h180_polygon.area
                                logger.info(f"Zone h<1.80m {h180_id} valide trouvée avec surface: {h180_area}")
                                
                                # Journaliser la forme de la zone h<1.80m
                                try:
                                    h180_coords = mapping(h180_polygon)
                                    logger.info(f"Forme de la zone h<1.80m {h180_id}: {h180_polygon.geom_type}")
                                except Exception as e:
                                    logger.warning(f"Impossible d'extraire les coordonnées de la zone h<1.80m {h180_id}: {str(e)}")
                            else:
                                logger.warning(f"Polygone de zone h<1.80m {h180_id} invalide après tentative de réparation")
                                continue
                        except Exception as e:
                            logger.warning(f"Erreur lors de la création du polygone h<1.80m {h180_id}: {str(e)}")
                            continue
                        
                        # Calculer l'intersection avec les polylignes de cette destination
                        intersection_found = False
                        destination_polylines = [p for p in main_projet_polylines if get_destination_from_layer(p.get('layer', '')) == destination]
                        
                        logger.info(f"Vérification des intersections pour zone h<1.80m {h180_id} avec {destination} ({len(destination_polylines)} polylignes)")
                        
                        # Essayer d'abord avec un buffer légèrement plus grand pour capturer les intersections proches
                        h180_polygon_buffered = h180_polygon.buffer(0.05)  # Augmenter le buffer à 0.05 pour mieux capturer les intersections
                        
                        for main_polyline in destination_polylines:
                            main_id = main_polyline.get('id', 'inconnu')
                            main_layer = main_polyline.get('layer', 'inconnu')
                            logger.info(f"Tentative d'intersection entre zone h<1.80m {h180_id} et polyligne {main_id} (calque {main_layer})")
                            
                            # Calculer l'aire d'intersection avec méthode améliorée
                            intersection_area = calculate_intersection_area(polyline, main_polyline)
                            
                            if intersection_area > 0:
                                surfaces_h_moins_180 += intersection_area
                                intersection_found = True
                                logger.info(f"Intersection trouvée entre zone h<1.80m {h180_id} et {destination} (polyligne {main_id}): {intersection_area}")
                            else:
                                # Essayer avec le buffer si l'intersection directe échoue
                                try:
                                    main_vertices = main_polyline.get('vertices', [])
                                    if len(main_vertices) >= 3:
                                        main_pts = [(float(v['x']), float(v['y'])) for v in main_vertices]
                                        if main_pts[0] != main_pts[-1]:
                                            main_pts.append(main_pts[0])
                                            
                                        main_polygon = Polygon(main_pts)
                                        if main_polygon.is_valid:
                                            # Vérifier l'intersection avec le buffer
                                            if h180_polygon_buffered.intersects(main_polygon):
                                                buffer_intersection = h180_polygon_buffered.intersection(main_polygon)
                                                buffer_area = buffer_intersection.area
                                                logger.info(f"Intersection avec buffer trouvée pour h<1.80m: {buffer_area}")
                                                
                                                # Utiliser l'aire de la zone h<1.80m originale si l'intersection est significative
                                                if buffer_area > 0.001:  # Seuil abaissé pour capturer plus d'intersections
                                                    # Si l'intersection directe est nulle, utiliser l'aire de la zone h<1.80m
                                                    # dans la zone d'intersection du buffer
                                                    intersection_ratio = buffer_area / h180_polygon_buffered.area
                                                    estimated_area = h180_area * intersection_ratio
                                                    
                                                    # Vérifier si l'estimation est raisonnable
                                                    if estimated_area > 0.001 and estimated_area <= h180_area:
                                                        surfaces_h_moins_180 += estimated_area
                                                        intersection_found = True
                                                        logger.info(f"Intersection estimée entre zone h<1.80m {h180_id} et {destination} (polyligne {main_id}): {estimated_area}")
                                                    else:
                                                        # Essayer une dernière fois avec un calcul direct
                                                        actual_intersection = h180_polygon.intersection(main_polygon)
                                                        actual_area = actual_intersection.area
                                                        if actual_area > 0:
                                                            surfaces_h_moins_180 += actual_area
                                                            intersection_found = True
                                                            logger.info(f"Intersection directe après buffer entre zone h<1.80m {h180_id} et {destination} (polyligne {main_id}): {actual_area}")
                                                        elif buffer_area > 0.1:  # Si l'intersection avec buffer est significative
                                                            # Utiliser une valeur minimale basée sur le buffer
                                                            min_area = min(h180_area, buffer_area * 0.5)
                                                            surfaces_h_moins_180 += min_area
                                                            intersection_found = True
                                                            logger.info(f"Intersection minimale entre zone h<1.80m {h180_id} et {destination} (polyligne {main_id}): {min_area}")
                                                            logger.info(f"  - Aire de la zone h<1.80m: {h180_area}")
                                                            logger.info(f"  - Aire de l'intersection avec buffer: {buffer_area}")
                                                            logger.info(f"  - Aire minimale utilisée: {min_area}")
                                                            logger.info(f"  - Distance entre polygones: {h180_polygon.distance(main_polygon)}")
                                                            
                                                            # Journaliser la valeur finale des surfaces h<1.80m pour cette destination
                                                            logger.info(f"VALEUR FINALE DES SURFACES H<1.80m POUR {destination}: {surfaces_h_moins_180}")
                                                            
                                                            # Forcer la mise à jour dans ta_data pour cette destination
                                                            if destination in ta_data:
                                                                ta_data[destination]['surfaces_h_moins_180'] = surfaces_h_moins_180
                                except Exception as e:
                                    logger.warning(f"Erreur lors de la vérification d'intersection avec buffer pour h<1.80m: {str(e)}")
                        
                        # Journaliser si aucune intersection n'a été trouvée pour cette zone h<1.80m avec cette destination
                        if not intersection_found:
                            logger.warning(f"Aucune intersection trouvée entre la zone h<1.80m {h180_id} ({h180_area}) et la destination {destination}")
                
                # Calculer le TOTAL T.A.
                total_ta = planchers_avant_deductions - (vides + surfaces_h_moins_180)
                
                # S'assurer que le total n'est pas négatif
                total_ta = max(0, total_ta)
                
                # Journaliser les valeurs calculées avant de les stocker
                logger.info(f"DESTINATION {destination} - VALEURS FINALES:")
                logger.info(f"  - Planchers avant déductions: {planchers_avant_deductions}")
                logger.info(f"  - Vides: {vides}")
                logger.info(f"  - Surfaces h<1.80m: {surfaces_h_moins_180}")
                logger.info(f"  - Total TA: {total_ta}")
                
                # Stocker les données pour cette destination
                ta_data[destination] = {
                    'planchers_avant_deductions': planchers_avant_deductions,
                    'vides': vides,
                    'surfaces_h_moins_180': surfaces_h_moins_180,
                    'total_ta': total_ta
                }
            
            # Ajouter les données T.A. à la feuille TA
            ta_row = 2
            for destination in sorted(all_destinations):
                # Formater le nom de la destination pour une meilleure lisibilité
                formatted_destination = special_cases.get(destination, destination.replace('_', ' ').lower().capitalize())
                
                # Colonne B: Destination
                ws_ta[f'B{ta_row}'] = formatted_destination
                cell = ws_ta[f'B{ta_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Récupérer les données T.A. pour cette destination
                destination_ta_data = ta_data.get(destination, {})
                
                # Colonne C: TA avant déduction (renommé de "Planchers avant déductions")
                planchers_avant_deductions = destination_ta_data.get('planchers_avant_deductions', 0)
                ws_ta[f'C{ta_row}'] = round(planchers_avant_deductions, 4) if planchers_avant_deductions > 0 else 0
                cell = ws_ta[f'C{ta_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Colonne D: Vides
                vides = destination_ta_data.get('vides', 0)
                logger.info(f"EXCEL - {formatted_destination} - Vides: {vides}")
                excel_vides_value = round(vides, 4) if vides > 0 else 0
                ws_ta[f'D{ta_row}'] = excel_vides_value
                logger.info(f"EXCEL - Valeur écrite dans D{ta_row}: {excel_vides_value}")
                cell = ws_ta[f'D{ta_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Colonne E: Surfaces dont h < 1.80m
                surfaces_h_moins_180 = destination_ta_data.get('surfaces_h_moins_180', 0)
                logger.info(f"EXCEL - {formatted_destination} - Surfaces h<1.80m: {surfaces_h_moins_180}")
                excel_h180_value = round(surfaces_h_moins_180, 4) if surfaces_h_moins_180 > 0 else 0
                ws_ta[f'E{ta_row}'] = excel_h180_value
                logger.info(f"EXCEL - Valeur écrite dans E{ta_row}: {excel_h180_value}")
                cell = ws_ta[f'E{ta_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Colonne F: TA après déduction (renommé de "TOTAL T.A.")
                total_ta = destination_ta_data.get('total_ta', 0)
                ws_ta[f'F{ta_row}'] = round(total_ta, 4) if total_ta > 0 else 0
                cell = ws_ta[f'F{ta_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Ajouter un commentaire expliquant le calcul
                comment = Comment(f"TA après déduction = TA avant déduction - (Vides + Surfaces dont h < 1.80m)", "Calcul T.A.")
                comment.width = 300
                comment.height = 50
                cell.comment = comment
                
                ta_row += 1
            
            # Nous n'ajoutons plus les colonnes SDP au tableau principal car elles sont maintenant dans la feuille SDP dédiée
            
            # Fonction pour déterminer si une polyligne est un calque spécifique
            def is_specific_layer_polyline(polyline, layer_pattern):
                layer = polyline.get('layer', '')
                if not isinstance(layer, str):
                    return False
                return layer_pattern in layer
            
            # Fonction pour calculer la surface d'une polyligne
            def calculate_polyline_area(polyline):
                area = 0.0
                vertices = polyline.get('vertices', [])
                if len(vertices) >= 3:
                    # Calculer la surface en utilisant la formule de Shoelace (Gauss's area formula)
                    n = len(vertices)
                    for i in range(n):
                        j = (i + 1) % n
                        area += vertices[i]['x'] * vertices[j]['y']
                        area -= vertices[j]['x'] * vertices[i]['y']
                    area = abs(area) / 2.0
                return area
            
            # Remplir la feuille SDP avec les données des fichiers DXF
            sdp_row = 2
            for destination in sorted(all_destinations):
                # Formater le nom de la destination pour une meilleure lisibilité
                formatted_destination = special_cases.get(destination, destination)
                
                # Récupérer les données pour cette destination depuis les résultats de calcul
                # Directement depuis les fichiers DXF traités:
                # - Projet_demoli_feuille_TA.dxf -> surfaces existantes
                # - Existant_exmple_demoli.dxf -> surfaces projet
                
                # Surface existante avant travaux (A) - directement depuis Projet_demoli_feuille_TA.dxf
                surface_existante = calculation_results['existant'].get(destination, 0)
                
                # Surface projet - directement depuis Existant_exmple_demoli.dxf
                surface_projet = calculation_results['projet'].get(destination, 0)
                
                # Surface créée (B) - différence entre projet et existant (si positive)
                surface_creee = max(0, surface_projet - surface_existante)
                
                # Surface démolie reconstruite - zones démolies qui sont reconstruites
                # Ces zones sont identifiées par l'intersection des polylignes de démolition avec les polylignes existantes
                surface_demolie_reconstruite = calculation_results['demolition'].get(destination, 0)
                
                # Surface supprimée (D) - zones existantes qui ne sont plus présentes dans le projet
                # et qui ne sont pas reconstruites
                surface_supprimee = max(0, surface_existante - surface_projet - surface_demolie_reconstruite)
                if surface_supprimee < 0.01:  # Éviter les valeurs négligeables dues aux erreurs d'arrondi
                    surface_supprimee = 0
                
                # Surface supprimée par changement de destination
                # Dans ce cas, nous utilisons directement les données calculées
                surface_supprimee_changement = calculation_results['supprimee_changement'].get(destination, 0)
                
                # Surface RDV (Réglementation de Voirie)
                # Selon les règles métier, la surface RDV est inférieure à la surface projet
                # Elle est calculée en fonction de la destination
                if 'HABITATION' in destination.upper():
                    surface_rdv = surface_projet * 0.75  # 75% pour l'habitation
                elif 'COMMERCE' in destination.upper():
                    surface_rdv = surface_projet * 0.80  # 80% pour le commerce
                else:
                    surface_rdv = surface_projet * 0.85  # 85% pour les autres destinations
                
                # Ajouter une ligne dans la feuille SDP
                # Write floor name only in the first row
                if sdp_row == 2:  # First row of data
                    ws_sdp[f'A{sdp_row}'] = floor_name
                else:
                    ws_sdp[f'A{sdp_row}'] = None  # Empty cell for subsequent rows
                ws_sdp[f'B{sdp_row}'] = formatted_destination
                
                # Remplir les colonnes avec les données calculées
                ws_sdp[f'C{sdp_row}'] = round(surface_existante, 4) if surface_existante > 0 else ""
                ws_sdp[f'D{sdp_row}'] = round(surface_creee, 4) if surface_creee > 0 else ""
                ws_sdp[f'E{sdp_row}'] = round(surface_demolie_reconstruite, 4) if surface_demolie_reconstruite > 0 else ""
                ws_sdp[f'F{sdp_row}'] = round(surface_supprimee, 4) if surface_supprimee > 0 else ""
                ws_sdp[f'G{sdp_row}'] = round(surface_supprimee_changement, 4) if surface_supprimee_changement > 0 else ""
                ws_sdp[f'H{sdp_row}'] = round(surface_projet, 4) if surface_projet > 0 else ""
                ws_sdp[f'I{sdp_row}'] = round(surface_rdv, 4) if surface_rdv > 0 else ""
                
                # Appliquer le style aux cellules
                data_font = Font(name='Arial', size=12)
                data_alignment = Alignment(horizontal='center', vertical='center')
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    cell = ws_sdp[f'{col}{sdp_row}']
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                
                sdp_row += 1
            
            # Créer une nouvelle feuille nommée "TA Existant" avec la même structure que "TA Projet"
            ws_ta_existant = wb.create_sheet(title="TA Existant")
            
            # En-têtes pour la feuille TA Existant (même structure que TA Projet)
            ws_ta_existant['A1'] = "Étages"
            ws_ta_existant['B1'] = "Destinations"
            ws_ta_existant['C1'] = "TA avant déduction"
            ws_ta_existant['D1'] = "Vides"
            ws_ta_existant['E1'] = "Surfaces dont h < 1.80m"
            ws_ta_existant['F1'] = "TA après déduction"
            
            # Appliquer le style aux en-têtes
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws_ta_existant[f'{col}1']
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws_ta_existant.column_dimensions[col].width = 15  # Ajuster la largeur des colonnes
            
            # Traiter les données pour la feuille "TA Existant" (basé sur existant.dxf)
            # Nous utilisons les polylignes existantes qui sont dans le fichier existant.dxf
            # Filtrer les polylignes principales et spéciales pour existant.dxf
            main_existant_polylines = [p for p in existant_polylines if is_main_sdp_polyline(p)]
            special_existant_polylines = [p for p in existant_polylines if is_special_polyline(p)]
            
            # Collecter toutes les destinations des polylignes principales
            all_existant_destinations = set()
            for polyline in main_existant_polylines:
                destination = get_destination_from_layer(polyline.get('layer', ''))
                if destination:
                    all_existant_destinations.add(destination)
            
            # Dictionnaire pour stocker les données T.A. Existant par destination
            ta_existant_data = {}
            
            # Calculer les données T.A. Existant pour chaque destination
            for destination in all_existant_destinations:
                # Initialiser les valeurs pour cette destination
                planchers_avant_deductions = 0.0  # Planchers avant déductions
                vides = 0.0                       # Vides (stairwells, elevator shafts, etc.)
                surfaces_h_moins_180 = 0.0        # Surfaces dont h < 1.80m
                
                # Calculer Planchers avant déductions - utiliser les surfaces principales
                for polyline in main_existant_polylines:
                    if get_destination_from_layer(polyline.get('layer', '')) == destination:
                        # Calculer la surface de cette polyligne
                        area = calculate_polyline_area(polyline)
                        planchers_avant_deductions += area
                
                # Identifier les vides (GEX_EDS_SDP_2 - TREMIE)
                for polyline in special_existant_polylines:
                    special_layer = polyline.get('layer', '')
                    if not isinstance(special_layer, str):
                        continue
                    
                    if 'GEX_EDS_SDP_2' in special_layer:  # Vides/TREMIE
                        # Calculer l'intersection avec les polylignes de cette destination
                        destination_polylines = [p for p in main_existant_polylines if get_destination_from_layer(p.get('layer', '')) == destination]
                        
                        for main_polyline in destination_polylines:
                            # Calculer l'aire d'intersection avec méthode améliorée
                            intersection_area = calculate_intersection_area(polyline, main_polyline)
                            if intersection_area > 0:
                                vides += intersection_area
                
                # Identifier les surfaces dont h < 1.80m (GEX_EDS_SDP_3 - H-180)
                for polyline in special_existant_polylines:
                    special_layer = polyline.get('layer', '')
                    if not isinstance(special_layer, str):
                        continue
                    
                    if 'GEX_EDS_SDP_3' in special_layer:  # H-180
                        # Calculer l'intersection avec les polylignes de cette destination
                        destination_polylines = [p for p in main_existant_polylines if get_destination_from_layer(p.get('layer', '')) == destination]
                        
                        for main_polyline in destination_polylines:
                            # Calculer l'aire d'intersection avec méthode améliorée
                            intersection_area = calculate_intersection_area(polyline, main_polyline)
                            if intersection_area > 0:
                                surfaces_h_moins_180 += intersection_area
                
                # Calculer le TOTAL T.A.
                total_ta = planchers_avant_deductions - (vides + surfaces_h_moins_180)
                
                # S'assurer que le total n'est pas négatif
                total_ta = max(0, total_ta)
                
                # Stocker les données pour cette destination
                ta_existant_data[destination] = {
                    'planchers_avant_deductions': planchers_avant_deductions,
                    'vides': vides,
                    'surfaces_h_moins_180': surfaces_h_moins_180,
                    'total_ta': total_ta
                }
            
            # Ajouter les données T.A. Existant à la feuille TA Existant
            ta_existant_row = 2
            ws_ta_existant[f'A{ta_existant_row}'] = floor_name
            
            # Appliquer le style à la cellule de l'étage
            cell = ws_ta_existant[f'A{ta_existant_row}']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            for destination in sorted(all_existant_destinations):
                # Formater le nom de la destination pour une meilleure lisibilité
                formatted_destination = special_cases.get(destination, destination.replace('_', ' ').lower().capitalize())
                
                # Colonne B: Destination
                ws_ta_existant[f'B{ta_existant_row}'] = formatted_destination
                cell = ws_ta_existant[f'B{ta_existant_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Récupérer les données T.A. pour cette destination
                destination_ta_data = ta_existant_data.get(destination, {})
                
                # Colonne C: TA avant déduction
                planchers_avant_deductions = destination_ta_data.get('planchers_avant_deductions', 0)
                ws_ta_existant[f'C{ta_existant_row}'] = round(planchers_avant_deductions, 4) if planchers_avant_deductions > 0 else 0
                cell = ws_ta_existant[f'C{ta_existant_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Colonne D: Vides
                vides = destination_ta_data.get('vides', 0)
                excel_vides_value = round(vides, 4) if vides > 0 else 0
                ws_ta_existant[f'D{ta_existant_row}'] = excel_vides_value
                cell = ws_ta_existant[f'D{ta_existant_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Colonne E: Surfaces dont h < 1.80m
                surfaces_h_moins_180 = destination_ta_data.get('surfaces_h_moins_180', 0)
                excel_h180_value = round(surfaces_h_moins_180, 4) if surfaces_h_moins_180 > 0 else 0
                ws_ta_existant[f'E{ta_existant_row}'] = excel_h180_value
                cell = ws_ta_existant[f'E{ta_existant_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Colonne F: TA après déduction
                total_ta = destination_ta_data.get('total_ta', 0)
                ws_ta_existant[f'F{ta_existant_row}'] = round(total_ta, 4) if total_ta > 0 else 0
                cell = ws_ta_existant[f'F{ta_existant_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Ajouter un commentaire expliquant le calcul
                comment = Comment(f"TA après déduction = TA avant déduction - (Vides + Surfaces dont h < 1.80m)", "Calcul T.A.")
                comment.width = 300
                comment.height = 50
                cell.comment = comment
                
                ta_existant_row += 1
            
            # Créer une nouvelle feuille nommée "TA Summary"
            ws_ta_summary = wb.create_sheet(title="TA Summary")
            
            # En-têtes pour la feuille TA Summary avec les nouvelles colonnes demandées
            ws_ta_summary['A1'] = "Étages"
            ws_ta_summary['B1'] = "TA Existant"
            ws_ta_summary['C1'] = "TA Projet"
            ws_ta_summary['D1'] = "TA créé"
            ws_ta_summary['E1'] = "TA démoli/reconstruit"
            ws_ta_summary['F1'] = "TA supprimé"
            
            # Appliquer le style aux en-têtes
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws_ta_summary[f'{col}1']
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws_ta_summary.column_dimensions[col].width = 18  # Ajuster la largeur des colonnes
            
            # Calculer les totaux pour le résumé
            # Calculer le total de TA Existant à partir de la feuille TA Existant
            ta_existant_total = 0
            for destination in all_existant_destinations:
                destination_ta_data = ta_existant_data.get(destination, {})
                total_ta = destination_ta_data.get('total_ta', 0)
                ta_existant_total += total_ta
            
            # Calculer le total de TA Projet à partir de la feuille TA Projet
            ta_projet_total = 0
            for destination in all_destinations:
                destination_ta_data = ta_data.get(destination, {})
                total_ta = destination_ta_data.get('total_ta', 0)
                ta_projet_total += total_ta
            
            # Calculer les autres métriques de comparaison
            # TA créé - nouvelles surfaces (différence positive entre projet et existant)
            ta_cree_total = max(0, ta_projet_total - ta_existant_total)
            
            # TA démoli/reconstruit - zones démolies qui sont reconstruites
            ta_demoli_reconstruit_total = 0
            for destination in all_destinations:
                ta_demoli_reconstruit = calculation_results['demolition'].get(destination, 0)
                ta_demoli_reconstruit_total += ta_demoli_reconstruit
            
            # TA supprimé - zones existantes qui ne sont plus présentes dans le projet
            # et qui ne sont pas reconstruites
            ta_supprime_total = max(0, ta_existant_total - ta_projet_total + ta_demoli_reconstruit_total)
            if ta_supprime_total < 0.01:  # Éviter les valeurs négligeables dues aux erreurs d'arrondi
                ta_supprime_total = 0
                
            # Journaliser les totaux calculés pour débogage
            logger.info(f"TA Existant total: {ta_existant_total}")
            logger.info(f"TA Projet total: {ta_projet_total}")
            logger.info(f"TA créé total: {ta_cree_total}")
            logger.info(f"TA démoli/reconstruit total: {ta_demoli_reconstruit_total}")
            logger.info(f"TA supprimé total: {ta_supprime_total}")
            
            # Ajouter les données à la feuille TA Summary
            ws_ta_summary['A2'] = floor_name
            ws_ta_summary['B2'] = round(ta_existant_total, 4) if ta_existant_total > 0 else 0
            ws_ta_summary['C2'] = round(ta_projet_total, 4) if ta_projet_total > 0 else 0
            ws_ta_summary['D2'] = round(ta_cree_total, 4) if ta_cree_total > 0 else 0
            ws_ta_summary['E2'] = round(ta_demoli_reconstruit_total, 4) if ta_demoli_reconstruit_total > 0 else 0
            ws_ta_summary['F2'] = round(ta_supprime_total, 4) if ta_supprime_total > 0 else 0
            
            # Ajouter des commentaires expliquant les calculs
            comment_cree = Comment(f"TA créé = TA Projet - TA Existant (si positif)", "Calcul TA créé")
            comment_cree.width = 300
            comment_cree.height = 50
            ws_ta_summary['D2'].comment = comment_cree
            
            comment_supprime = Comment(f"TA supprimé = TA Existant - TA Projet + TA démoli/reconstruit (si positif)", "Calcul TA supprimé")
            comment_supprime.width = 300
            comment_supprime.height = 50
            ws_ta_summary['F2'].comment = comment_supprime
            
            # Appliquer le style aux cellules
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws_ta_summary[f'{col}2']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
            
            # No cell merging, etage name only appears in the first row
            
            # Sauvegarder le fichier Excel
            wb.save(excel_path)
            
            logger.info(f"Fichier Excel créé avec succès: {excel_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la génération du fichier Excel: {str(e)}")
            return jsonify({'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'}), 500
        
        # Réponse finale
        return jsonify({
            'message': 'Fichier Excel généré avec succès',
            'filePath': excel_path
        }), 201
    
    except Exception as e:
        # Gestionnaire général d'exceptions - attrape tout et journalise
        import traceback
        logger.error(f"ERREUR CRITIQUE dans generate_excel_file: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'}), 500

def download_excel_file():
    """Permet le téléchargement d'un fichier Excel"""
    logger.info("Requête GET reçue pour télécharger un fichier Excel")
    
    try:
        file_path = request.args.get('filePath')
        email = request.args.get('email')
        
        if not file_path:
            logger.error("Chemin de fichier non fourni")
            return jsonify({'error': 'Chemin de fichier non fourni'}), 400
        
        if not email:
            logger.error("Email non fourni")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Fichier demandé: {file_path} pour l'utilisateur: {email}")
        
        # Extrait le nom du fichier du chemin (pour recherche directe)
        filename = os.path.basename(file_path)
        logger.info(f"Nom du fichier: {filename}")
        
        # Déterminer le dossier racine de l'utilisateur basé sur son email
        current_dir = os.path.dirname(os.path.abspath(__file__))
        logger.info(f"Répertoire courant: {current_dir}")
        
        folder_name = email.split('@')[0]
        
        # Essayons différentes combinaisons possibles pour le répertoire utilisateur
        possible_paths = [
            os.path.join(current_dir, 'app', 'Ressources', folder_name),
            os.path.join(current_dir, 'Ressources', folder_name),
        ]
        
        # Vérifier chaque chemin possible
        user_path_found = None
        for path in possible_paths:
            logger.info(f"Vérification du chemin: {path} (existe: {os.path.exists(path)})")
            if os.path.exists(path):
                user_path_found = path
                break
        
        if user_path_found is None:
            logger.error(f"Aucun dossier utilisateur trouvé pour {email}")
            return jsonify({'error': 'Dossier utilisateur non trouvé'}), 404
        
        # Premier essai: rechercher le fichier dans le chemin fourni
        full_file_path = os.path.join(user_path_found, file_path)
        logger.info(f"Tentative 1 - Chemin complet du fichier: {full_file_path}")
        
        # Si le fichier n'existe pas, rechercher récursivement dans le répertoire de l'utilisateur
        if not os.path.exists(full_file_path):
            logger.info("Recherche du fichier dans le répertoire utilisateur...")
            found_files = []
            
            # Recherche récursive du fichier
            for root, dirs, files in os.walk(user_path_found):
                if filename in files:
                    found_path = os.path.join(root, filename)
                    found_files.append(found_path)
                    logger.info(f"Fichier trouvé: {found_path}")
            
            if found_files:
                # Utiliser le premier fichier trouvé
                full_file_path = found_files[0]
                logger.info(f"Utilisation du fichier trouvé: {full_file_path}")
            else:
                logger.error(f"Aucun fichier {filename} trouvé dans le répertoire utilisateur")
                return jsonify({'error': 'Fichier non trouvé dans le répertoire utilisateur'}), 404
        
        if not os.path.exists(full_file_path):
            logger.error(f"Fichier non trouvé: {full_file_path}")
            return jsonify({'error': 'Fichier non trouvé'}), 404
            
        # Le fichier existe, vérifier sa taille
        file_size = os.path.getsize(full_file_path)
        logger.info(f"Taille du fichier: {file_size} octets")
        
        if file_size == 0:
            logger.error("Le fichier est vide")
            return jsonify({'error': 'Le fichier est vide'}), 500
        
        logger.info(f"Envoi du fichier: {full_file_path}")
        
        # Envoyer le fichier au client
        try:
            return send_file(
                full_file_path,
                as_attachment=True,
                download_name=os.path.basename(file_path),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi du fichier: {str(e)}")
            return jsonify({'error': f"Erreur lors de l'envoi du fichier: {str(e)}"}), 500
    
    except Exception as e:
        logger.error(f"Erreur lors du téléchargement du fichier Excel: {str(e)}")
        return jsonify({'error': f'Erreur lors du téléchargement du fichier: {str(e)}'}), 500

def generate_visa_content(surfaces, floor_name):
    """Génère le contenu du fichier visa.txt avec tous les détails des calculs et des éléments"""
    now = datetime.datetime.now()
    content = [
        "=" * 80,
        f"RAPPORT DÉTAILLÉ DE CALCUL DE SURFACE - {floor_name}",
        "=" * 80,
        f"Date de génération: {now.strftime('%d/%m/%Y %H:%M:%S')}",
        f"Étage/Niveau: {floor_name}",
        "",
        "RÉSUMÉ DES SURFACES:",
        "-" * 50
    ]
    
    projet_surface = surfaces.get('projet', {}).get('surface', 0)
    content.append(f"Surface PROJET:\t\t{projet_surface:.2f} m²")
    
    existant_surface = surfaces.get('existant', {}).get('surface', 0)
    content.append(f"Surface EXISTANT:\t{existant_surface:.2f} m²")
    
    difference = surfaces.get('difference', 0)
    if existant_surface > 0:
        perc_diff = difference / existant_surface * 100
        content.append(f"DIFFÉRENCE:\t\t{difference:.2f} m² ({'+' if difference > 0 else ''}{perc_diff:.2f}% par rapport à l'existant)")
    else:
        content.append(f"DIFFÉRENCE:\t\t{difference:.2f} m²")
    content.append(f"IMPACT:\t\t\t{'Agrandissement' if difference > 0 else 'Réduction' if difference < 0 else 'Pas de changement'}")
    
    if projet_surface > 0 and surfaces.get('projet', {}).get('details', {}).get('polylines', 0) > 0:
        densite_projet = projet_surface / surfaces['projet']['details']['polylines']
        content.append(f"DENSITÉ PROJET:\t{densite_projet:.2f} m²/élément")
    if existant_surface > 0 and surfaces.get('existant', {}).get('details', {}).get('polylines', 0) > 0:
        densite_existant = existant_surface / surfaces['existant']['details']['polylines']
        content.append(f"DENSITÉ EXISTANT:\t{densite_existant:.2f} m²/élément")
    
    content.append("")
    content.append("ANALYSE COMPARATIVE DÉTAILLÉE:")  
    content.append("-" * 50)
    
    if projet_surface > 0 and existant_surface > 0:
        if difference > 0:
            content.append(f"Détails de l'agrandissement: {difference:.2f} m² ajoutés par rapport à l'existant")
            content.append(f"Ratio surface projet/existant: {projet_surface/existant_surface:.2f}")
            content.append(f"Pourcentage d'augmentation: {perc_diff:.2f}%")
            proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
            exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
            if proj_elems > exist_elems:
                content.append(f"Surface moyenne ajoutée par nouvel élément: {difference / (proj_elems - exist_elems):.2f} m²")
            
        elif difference < 0:
            content.append(f"Détails de la réduction: {abs(difference):.2f} m² supprimés par rapport à l'existant")
            content.append(f"Ratio surface projet/existant: {projet_surface/existant_surface:.2f}")
            content.append(f"Pourcentage de diminution: {abs(perc_diff):.2f}%")
            proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
            exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
            if exist_elems > proj_elems:
                content.append(f"Surface moyenne réduite par élément supprimé: {abs(difference) / (exist_elems - proj_elems):.2f} m²")
        else:
            content.append("Les surfaces projet et existant sont identiques")
            content.append("Aucune modification significative de la surface globale")
            proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
            exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
            if proj_elems != exist_elems:
                content.append(f"Bien que la surface totale soit identique, le nombre d'éléments a changé: {proj_elems - exist_elems:+d} éléments")
    elif projet_surface > 0 and existant_surface == 0:
        content.append(f"Nouvelle construction: {projet_surface:.2f} m² sans existant précédent")
        proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
        if proj_elems > 0:
            content.append(f"Surface moyenne par élément: {projet_surface / proj_elems:.2f} m²")
    elif projet_surface == 0 and existant_surface > 0:
        content.append(f"Démolition complète: {existant_surface:.2f} m² de surface existante")
        exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
        if exist_elems > 0:
            content.append(f"Surface moyenne par élément supprimé: {existant_surface / exist_elems:.2f} m²")
        
    content.append("")
    content.append("INVENTAIRE DÉTAILLÉ DES ÉLÉMENTS:")
    content.append("-" * 50)
    
    if surfaces.get('projet', {}).get('details'):
        projet_details = surfaces['projet']['details']
        content.append("PROJET:")
        content.append(f"  - Nombre total de polylignes: {projet_details.get('polylines', 0)}")
        content.append(f"  - Nombre total de cercles: {projet_details.get('circles', 0)}")
        content.append(f"  - Nombre total d'éléments: {projet_details.get('polylines', 0) + projet_details.get('circles', 0)}")
        if 'polylines' in projet_details and projet_details['polylines'] > 0:
            content.append(f"  - Surface moyenne par polyligne: {projet_surface / projet_details['polylines']:.2f} m² (si répartition uniforme)")
        
        content.append("  - Types d'éléments: Polylignes fermées (surfaces), cercles")
        if projet_details.get('polylines', 0) > 0 and projet_details.get('circles', 0) > 0:
            content.append(f"  - Ratio polylignes/cercles: {projet_details.get('polylines', 0) / projet_details.get('circles', 1):.2f}")
        
    if surfaces.get('existant', {}).get('details'):
        existant_details = surfaces['existant']['details']
        content.append("\nEXISTANT:")
        content.append(f"  - Nombre total de polylignes: {existant_details.get('polylines', 0)}")
        content.append(f"  - Nombre total de cercles: {existant_details.get('circles', 0)}")
        content.append(f"  - Nombre total d'éléments: {existant_details.get('polylines', 0) + existant_details.get('circles', 0)}")
        if 'polylines' in existant_details and existant_details['polylines'] > 0:
            content.append(f"  - Surface moyenne par polyligne: {existant_surface / existant_details['polylines']:.2f} m² (si répartition uniforme)")
        
        content.append("  - Types d'éléments: Polylignes fermées (surfaces), cercles")
        if existant_details.get('polylines', 0) > 0 and existant_details.get('circles', 0) > 0:
            content.append(f"  - Ratio polylignes/cercles: {existant_details.get('polylines', 0) / existant_details.get('circles', 1):.2f}")
    
    if surfaces.get('projet', {}).get('details') and surfaces.get('existant', {}).get('details'):
        projet_details = surfaces['projet']['details']
        existant_details = surfaces['existant']['details']
        
        content.append("\nCOMPARAISON DES ÉLÉMENTS:")
        poly_diff = projet_details.get('polylines', 0) - existant_details.get('polylines', 0)
        circle_diff = projet_details.get('circles', 0) - existant_details.get('circles', 0)
        total_elem_diff = poly_diff + circle_diff
        
        content.append(f"  - Différence en polylignes: {'+' if poly_diff > 0 else ''}{poly_diff}")
        content.append(f"  - Différence en cercles: {'+' if circle_diff > 0 else ''}{circle_diff}")
        content.append(f"  - Différence totale d'éléments: {'+' if total_elem_diff > 0 else ''}{total_elem_diff}")
    
    content.append("")
    content.append("COMMENTAIRES ET OBSERVATIONS:")
    content.append("-" * 50)
    
    if difference > 0:
        content.append(f"- Ce calcul montre une augmentation de surface de {difference:.2f} m² ({perc_diff:.2f}% par rapport à l'existant).")
        content.append("- Veuillez vérifier que cette augmentation est conforme aux règles d'urbanisme et aux limites de constructibilité.")
        content.append(f"- Vérification recommandée: coefficient d'emprise au sol (CES) et coefficient d'occupation des sols (COS) du PLU.")
        
        content.append("- Points d'attention:")
        content.append("  * Vérifier les calculs des droits à construire dans le cas d'une extension")
        content.append("  * Contrôler la conformité avec les règles de constructibilité locales")
        content.append("  * Confirmer la compatibilité avec les règles de gabarit et de prospect")
        
        if perc_diff > 20:
            content.append(f"- ATTENTION: L'augmentation de surface de {perc_diff:.2f}% est significative et peut nécessiter ")
            content.append("  des autorisations d'urbanisme spécifiques (permis de construire plutôt qu'une déclaration préalable).")
        
    elif difference < 0:
        content.append(f"- Ce calcul montre une diminution de surface de {abs(difference):.2f} m² ({abs(perc_diff):.2f}% par rapport à l'existant).")
        content.append("- Assurez-vous que cette réduction est voulue et qu'elle respecte les objectifs du projet.")
        
        content.append("- Points d'attention:")
        content.append("  * Vérifier l'impact sur le fonctionnement des espaces")
        content.append("  * Contrôler la conformité des nouveaux espaces avec les normes d'accessibilité")
        content.append("  * Évaluer l'impact sur les calculs réglementaires (surface utile, surface taxable)")
        
        if abs(perc_diff) > 30:
            content.append(f"- NOTE: La réduction importante de {abs(perc_diff):.2f}% peut indiquer une restructuration majeure")
            content.append("  du bâtiment. Vérifiez l'exactitude des fichiers DXF et des éléments pris en compte.")
    else:
        content.append("- Les surfaces projet et existant sont identiques.")
        content.append("- Aucun impact sur la surface totale.")
        
        proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
        exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
        if proj_elems != exist_elems:
            content.append(f"- Bien que la surface totale soit inchangée, la répartition des espaces a été modifiée")
            content.append("  (nombre d'éléments différent entre projet et existant).")
            content.append("- Vérifier que cette redistribution respecte les exigences fonctionnelles du bâtiment.")
        else:
            content.append("- La structure du bâtiment semble inchangée (même nombre d'éléments).")
            content.append("- Il peut s'agir d'une mise à jour du fichier sans modification structurelle.")
    
    content.append("")
    content.append("RECOMMANDATIONS TECHNIQUES:")
    content.append("-" * 50)
    content.append("- Vérification de la cohérence des surfaces avec les autres documents du projet")
    content.append("- Confirmation des calculs avec les surfaces mentionnées dans les dossiers administratifs")
    content.append("- Évaluation de l'impact des modifications sur les performances énergétiques du bâtiment")
    content.append("- Analyse des implications sur les évacuations et issues de secours (si applicable)")
    
    content.append("")
    content.append("=" * 80)
    content.append("INFORMATIONS LÉGALES:")
    content.append("-" * 50)
    content.append(f"Ce rapport a été généré automatiquement le {now.strftime('%d/%m/%Y')} à {now.strftime('%H:%M:%S')}")
    content.append("Les calculs sont effectués sur la base des éléments fournis dans les fichiers DXF.")
    content.append("Ces résultats doivent être vérifiés par un professionnel qualifié.")
    content.append("")
    return "\n".join(content)

def create_excel_document(excel_path, surfaces, floor_name):
    """Crée un document Excel avec les données de surface"""
    wb = openpyxl.Workbook()
    
    # Configuration des styles communs
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    data_font = Font(name='Arial', size=12)
    
    title_alignment = Alignment(horizontal='center', vertical='center')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Création de la feuille TA (renommée en "Detailled TA")
    ws_ta = wb.active
    ws_ta.title = "Detailled TA"
    
    # En-têtes colonnes TA
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for col in columns:
        ws_ta.column_dimensions[col].width = 15
    
    # En-tête de la feuille TA avec explication des calculs de vide
    ws_ta['A1'] = "Étages"
    ws_ta['B1'] = "Destinations"
    ws_ta['C1'] = "TA avant déduction"
    ws_ta['D1'] = "Vides - Intersection précise"
    ws_ta['E1'] = "Surfaces dont h < 1.80m"
    ws_ta['F1'] = "TA après déduction"
    ws_ta['G1'] = "% Vide"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws_ta[f'{col}1']
        cell.font = subheader_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajouter uniquement le nom d'étage dans la feuille TA
    ws_ta['A2'] = floor_name
    
    # Appliquer le style uniquement à la cellule du nom d'étage
    cell = ws_ta['A2']
    cell.font = data_font
    cell.alignment = data_alignment
    cell.border = border
    
    # Création de la feuille SDP avec un nouveau titre pour tester
    ws_sdp = wb.create_sheet(title="SDP_VERIFICATION")
    
    # Configuration des colonnes SDP
    for col in columns:
        ws_sdp.column_dimensions[col].width = 15
    
    # En-tête de la feuille SDP - avec les titres modifiés et les colonnes divisées
    ws_sdp['A1'] = "Etages_TEST"
    ws_sdp['B1'] = "Destinations"
    ws_sdp['C1'] = "Surface existante avant travaux (A)"
    ws_sdp['D1'] = "Surface creee (B)"
    ws_sdp['E1'] = "Surface creee par changement de destination (C)"
    ws_sdp['F1'] = "Surface demolie reconstruite"
    ws_sdp['G1'] = "Surface supprimee (D)"
    ws_sdp['H1'] = "Surface supprimee par changement de destination"
    ws_sdp['I1'] = "Surface projet"
    ws_sdp['J1'] = "Surface RDV"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        cell = ws_sdp[f'{col}1']
        cell.font = subheader_font
        cell.alignment = header_alignment
        cell.border = border
        
    # Ajouter le nom d'étage dans la feuille SDP uniquement dans la première ligne
    ws_sdp['A2'] = floor_name
    
    # Appliquer le style à la cellule du nom d'étage
    cell = ws_sdp[f'A2']
    cell.font = data_font
    cell.alignment = data_alignment
    cell.border = border
    
    # Préparer la colonne A pour éviter la répétition du nom d'étage
    # Nous allons utiliser une formule pour laisser les cellules vides
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="custom", formula1="TRUE", showErrorMessage=False)
    ws_sdp.add_data_validation(dv)
    
    # Utiliser les données réelles des polylignes pour la colonne Destinations
    row = 2  # Commencer à la ligne 2
    
    # Fonction pour extraire le nom de destination à partir d'une polyligne
    def get_destination_name(polyline):
        import re
        
        # Essayer d'extraire le nom du calque (layer)
        layer = polyline.get('layer', '')
        
        # Si le calque n'est pas valide, essayer d'autres propriétés
        if not layer or not isinstance(layer, str):
            # Essayer d'utiliser le type de polyligne
            polyline_type = polyline.get('type', '')
            if polyline_type:
                return polyline_type
            return "NON_SPECIFIE"
        
        # Format standard: GEX_EDS_SDP_X-CATEGORIE_SOUSCATEGORIE
        # Exemple: GEX_EDS_SDP_1-COMMERCE_CIN -> COMMERCE_CIN
        pattern1 = r'(?:GEX_EDS_)?SDP_\d+-([A-Z_]+)'
        match = re.search(pattern1, layer)
        if match:
            # Extraire uniquement la partie après le tiret (COMMERCE_CIN)
            return match.group(1)
        
        # Format spécifique pour hauteur sous plafond (H-180)
        if "H-" in layer:
            match = re.search(r'H-(\d+)', layer)
            if match:
                return f"H-{match.group(1)}"
        
        # Format spécifique pour stationnement
        if "PK" in layer:
            return "PK"
        
        # Si le calque contient un tiret, prendre la partie après le tiret
        if "-" in layer:
            parts = layer.split('-')
            if len(parts) > 1:
                return parts[-1]
        
        # Enlever les préfixes communs
        clean_layer = re.sub(r'^(?:GEX_EDS_)?(?:SDP_)?\d*', '', layer)
        
        # Si le calque nettoyé commence par un tiret, l'enlever
        if clean_layer.startswith('-'):
            clean_layer = clean_layer[1:]
        
        # Si le calque nettoyé n'est pas vide, le retourner
        if clean_layer:
            return clean_layer
        
        # En dernier recours, retourner le calque original
        return layer
    
    # Collecter les noms de destination uniques et calculer les surfaces par destination
    destinations = set()  # Utiliser un ensemble pour éliminer automatiquement les doublons
    
    # Créer un dictionnaire pour stocker tous les types de calques rencontrés
    all_layer_types = {}
    
    # Fonction pour enregistrer un type de calque
    def register_layer_type(polyline):
        layer = polyline.get('layer', '')
        if layer and isinstance(layer, str):
            # Extraire le nom de destination
            destination = get_destination_name(polyline)
            all_layer_types[destination] = True
    
    # Dictionnaire pour stocker les surfaces par destination
    surfaces_par_destination = {}
    
    # Extraire les destinations des polylignes du projet
    projet_polylines = surfaces.get('projet', {}).get('polylines', [])
    for polyline in projet_polylines:
        destination = get_destination_name(polyline)
        destinations.add(destination)
    
    # Fonction pour calculer la surface d'une polyligne à partir de ses vertices
    def calculer_surface_polyligne(polyline):
        # Vérifier si la polyligne a des vertices
        if 'vertices' not in polyline or not polyline['vertices']:
            return 0
        
        vertices = polyline['vertices']
        if len(vertices) < 3:  # Un polygone doit avoir au moins 3 points
            return 0
        
        # Calculer la surface en utilisant la formule de Shoelace (Gauss's area formula)
        area = 0.0
        n = len(vertices)
        for i in range(n):
            j = (i + 1) % n
            area += vertices[i]['x'] * vertices[j]['y']
            area -= vertices[j]['x'] * vertices[i]['y']
        
        # Prendre la valeur absolue et diviser par 2
        return abs(area) / 2.0
    
    # IMPORTANT: Selon la demande du client et en tenant compte de l'inversion des fichiers
    # La colonne "Surface existante avant travaux (A)" contient les surfaces des fichiers de "Projet_demoli_feuille_TA.dxf" (section "Importation d'existant .dxf")
    # La colonne "Surface projet" contient les surfaces des fichiers de "Existant_exmple_demoli.dxf" (section "Importation de fichier .dxf")
    
    # Fonction pour déterminer si une polyligne est un calque principal SDP_1
    def is_main_sdp_polyline(polyline):
        layer = polyline.get('layer', '')
        return isinstance(layer, str) and layer.startswith('GEX_EDS_SDP_1')
    
    # Fonction pour déterminer si une polyligne est un calque spécial à déduire (SDP_2, SDP_3, etc.)
    def is_special_polyline(polyline):
        layer = polyline.get('layer', '')
        if not isinstance(layer, str):
            return False
        # Vérifier si c'est un calque spécial à déduire (pas SDP_1)
        patterns = ['GEX_EDS_SDP_2', 'GEX_EDS_SDP_3', 'GEX_EDS_SDP_4', 'GEX_EDS_SDP_5', 'GEX_EDS_SDP_7']
        return any(pattern in layer for pattern in patterns)
    
    # Fonction pour déterminer la destination parente d'un calque spécial de manière simple sans dépendre de Shapely
    def get_parent_destination(special_polyline, main_polylines):
        special_layer = special_polyline.get('layer', '')
        if not isinstance(special_layer, str):
            return None
            
        # Méthode simplifiée: utiliser les correspondances de noms de calques
        # Par exemple, si nous avons GEX_EDS_SDP_3-H-180, chercher un GEX_EDS_SDP_1 avec des coordonnées similaires
        special_prefix = special_layer.split('-')[0] if '-' in special_layer else ''
        
        # Essayer d'extraire le numéro d'étage ou de bâtiment (si présent dans le nom du calque)
        parts = special_prefix.split('_')
        etage_num = parts[-1] if len(parts) > 1 and parts[-1].isdigit() else None
        
        # Vérifier chaque polyligne principale pour voir si elle pourrait être parent
        # basé sur des heuristiques simples (proximité ou même étage)
        for main_polyline in main_polylines:
            main_layer = main_polyline.get('layer', '')
            if not isinstance(main_layer, str):
                continue
                
            # Si le calque spécial est inclus dans les coordonnées du calque principal, c'est probablement son parent
            if is_contained(special_polyline, main_polyline):
                return get_destination_name(main_polyline)
                
        # Si nous n'avons pas pu trouver de parent, retourner None
        return None
        
    # Fonction pour vérifier si une polyligne est contenue dans une autre polyligne
    # en utilisant une méthode simplifiée de comparaison de boîtes englobantes
    def is_contained(polyline1, polyline2):
        try:
            # Utiliser Shapely pour une vérification de contenance géométrique précise
            from shapely.geometry import Polygon
            
            # Créer les polygones à partir des sommets
            poly1_pts = [(v['x'], v['y']) for v in polyline1.get('vertices', [])]
            poly2_pts = [(v['x'], v['y']) for v in polyline2.get('vertices', [])]
            
            if len(poly1_pts) < 3 or len(poly2_pts) < 3:
                return False
                
            poly1_shapely = Polygon(poly1_pts)
            poly2_shapely = Polygon(poly2_pts)
            
            # Vérifier si les polygones sont valides
            if not poly1_shapely.is_valid or not poly2_shapely.is_valid:
                return False
            
            # Vérifier si poly1 est entièrement contenu dans poly2
            # Un polygone est contenu dans un autre si leur intersection est égale au premier polygone
            return poly2_shapely.contains(poly1_shapely)
        except Exception as e:
            logger.warning(f"Erreur lors de la vérification de contenance: {str(e)}")
            return False
            
    # Fonction pour vérifier si deux polylignes s'intersectent
    def intersects(polyline1, polyline2):
        try:
            from shapely.geometry import Polygon, box
            from shapely.validation import make_valid
            
            # Vérifier d'abord si les boîtes englobantes s'intersectent (rapide)
            poly1_pts = [(v['x'], v['y']) for v in polyline1.get('vertices', [])]
            poly2_pts = [(v['x'], v['y']) for v in polyline2.get('vertices', [])]
            
            if len(poly1_pts) < 3 or len(poly2_pts) < 3:
                return False
                
            # Vérification rapide des boîtes englobantes
            min_x1, min_y1 = min(p[0] for p in poly1_pts), min(p[1] for p in poly1_pts)
            max_x1, max_y1 = max(p[0] for p in poly1_pts), max(p[1] for p in poly1_pts)
            min_x2, min_y2 = min(p[0] for p in poly2_pts), min(p[1] for p in poly2_pts)
            max_x2, max_y2 = max(p[0] for p in poly2_pts), max(p[1] for p in poly2_pts)
            
            # Si les boîtes englobantes ne se chevauchent pas, retourner False
            if max_x1 < min_x2 or min_x1 > max_x2 or max_y1 < min_y2 or min_y1 > max_y2:
                return False
                
            # Si nous arrivons ici, les boîtes englobantes se chevauchent
            # Vérification spéciale pour congre_art
            dest1 = polyline1.get('destination', '')
            dest2 = polyline2.get('destination', '')
            
            if 'congre_art' in str(dest1).lower() or 'congre_art' in str(dest2).lower():
                # Pour congre_art, si les boîtes englobantes se chevauchent, considérer comme intersection
                logger.info(f"Intersection détectée par boîtes englobantes pour congre_art: {dest1}/{dest2}")
                return True
            
            # Pour les autres cas, effectuer une vérification géométrique précise
            try:
                poly1_shapely = Polygon(poly1_pts)
                poly2_shapely = Polygon(poly2_pts)
                
                # Réparer les polygones invalides
                if not poly1_shapely.is_valid:
                    poly1_shapely = make_valid(poly1_shapely)
                if not poly2_shapely.is_valid:
                    poly2_shapely = make_valid(poly2_shapely)
                
                # Si les polygones sont toujours invalides, utiliser l'intersection des boîtes englobantes
                if not poly1_shapely.is_valid or not poly2_shapely.is_valid:
                    logger.warning("Polygones invalides après tentative de réparation, utilisation des boîtes englobantes")
                    return True  # Les boîtes englobantes se chevauchent déjà, donc True
                
                # Vérifier s'ils s'intersectent
                intersects_result = poly1_shapely.intersects(poly2_shapely)
                
                # Si l'intersection échoue mais que les boîtes englobantes se chevauchent, essayer avec un petit buffer
                if not intersects_result:
                    buffer1 = poly1_shapely.buffer(0.01)
                    buffer2 = poly2_shapely.buffer(0.01)
                    if buffer1.intersects(buffer2):
                        logger.info(f"Intersection détectée avec buffer pour {dest1}/{dest2}")
                        return True
                
                return intersects_result
                
            except Exception as e:
                logger.warning(f"Erreur lors de la vérification géométrique d'intersection: {str(e)}")
                # En cas d'erreur, se rabattre sur l'intersection des boîtes englobantes
                return True  # Les boîtes englobantes se chevauchent déjà, donc True
            
        except Exception as e:
            logger.warning(f"Erreur lors de la vérification d'intersection: {str(e)}")
            return False
            
    # Fonction pour calculer l'intersection entre deux polylignes
    def calculate_intersection(poly1, poly2):
        try:
            # Convertir les sommets en objets Point
            from shapely.geometry import Polygon, box
            from shapely.validation import make_valid
            import numpy as np
            
            # Créer les polygones à partir des sommets
            poly1_pts = [(v['x'], v['y']) for v in poly1.get('vertices', [])]
            poly2_pts = [(v['x'], v['y']) for v in poly2.get('vertices', [])]
            
            if len(poly1_pts) < 3 or len(poly2_pts) < 3:
                return 0.0
            
            # Vérifier d'abord si les boîtes englobantes se chevauchent
            min_x1, min_y1 = min(p[0] for p in poly1_pts), min(p[1] for p in poly1_pts)
            max_x1, max_y1 = max(p[0] for p in poly1_pts), max(p[1] for p in poly1_pts)
            min_x2, min_y2 = min(p[0] for p in poly2_pts), min(p[1] for p in poly2_pts)
            max_x2, max_y2 = max(p[0] for p in poly2_pts), max(p[1] for p in poly2_pts)
            
            # Si les boîtes englobantes ne se chevauchent pas, retourner 0
            if max_x1 < min_x2 or min_x1 > max_x2 or max_y1 < min_y2 or min_y1 > max_y2:
                return 0.0
                
            try:
                poly1_shapely = Polygon(poly1_pts)
                poly2_shapely = Polygon(poly2_pts)
                
                # Réparer les polygones invalides
                if not poly1_shapely.is_valid:
                    poly1_shapely = make_valid(poly1_shapely)
                if not poly2_shapely.is_valid:
                    poly2_shapely = make_valid(poly2_shapely)
                
                # Si les polygones sont toujours invalides, essayer une approche alternative
                if not poly1_shapely.is_valid or not poly2_shapely.is_valid:
                    logger.warning("Polygones invalides après tentative de réparation, utilisation des boîtes englobantes")
                    box1 = box(min_x1, min_y1, max_x1, max_y1)
                    box2 = box(min_x2, min_y2, max_x2, max_y2)
                    intersection = box1.intersection(box2)
                    return intersection.area * 0.8  # Facteur de correction approximatif
                
                # Calculer l'intersection
                intersection = poly1_shapely.intersection(poly2_shapely)
                
                # Débogage pour les cas spéciaux (congre_art)
                dest1 = poly1.get('destination', '')
                dest2 = poly2.get('destination', '')
                layer1 = poly1.get('layer', '')
                layer2 = poly2.get('layer', '')
                
                if ('congre_art' in dest1 or 'congre_art' in dest2) and intersection.area == 0:
                    logger.warning(f"Intersection nulle détectée avec congre_art: {dest1} / {dest2}, layers: {layer1} / {layer2}")
                    # Essayer une approche alternative pour congre_art
                    buffer1 = poly1_shapely.buffer(0.01)  # Petit buffer pour aider à capturer les intersections
                    buffer2 = poly2_shapely.buffer(0.01)
                    alt_intersection = buffer1.intersection(buffer2)
                    if alt_intersection.area > 0:
                        logger.warning(f"Intersection alternative trouvée pour congre_art: {alt_intersection.area}")
                        return alt_intersection.area * 0.95  # Facteur de correction pour le buffer
                
                # Retourner l'aire de l'intersection
                return intersection.area
            except Exception as e:
                logger.warning(f"Erreur lors du calcul d'intersection avec Shapely: {str(e)}")
                # Approche de secours: utiliser les boîtes englobantes
                box1 = box(min_x1, min_y1, max_x1, max_y1)
                box2 = box(min_x2, min_y2, max_x2, max_y2)
                intersection = box1.intersection(box2)
                return intersection.area * 0.7  # Facteur de correction approximatif
            
        except Exception as e:
            logger.warning(f"Erreur lors du calcul d'intersection: {str(e)}")
            return 0.0
    
    # Récupérer les polylignes des deux fichiers
    existant_polylines = surfaces.get('existant', {}).get('polylines', [])
    projet_polylines = surfaces.get('projet', {}).get('polylines', [])
    
    # Initialiser le dictionnaire des surfaces
    surfaces_par_destination = {}
    
    # Filtrer les polylignes principales (SDP_1)
    main_existant_polylines = [p for p in existant_polylines if is_main_sdp_polyline(p)]
    main_projet_polylines = [p for p in projet_polylines if is_main_sdp_polyline(p)]
    
    # Filtrer les polylignes spéciales à déduire
    special_existant_polylines = [p for p in existant_polylines if is_special_polyline(p)]
    special_projet_polylines = [p for p in projet_polylines if is_special_polyline(p)]
    
    # Collecter toutes les destinations des polylignes principales
    all_destinations = set()
    for polyline in main_existant_polylines + main_projet_polylines:
        destination = get_destination_name(polyline)
        all_destinations.add(destination)
    
    # Initialiser le dictionnaire des surfaces pour toutes les destinations
    for destination in all_destinations:
        surfaces_par_destination[destination] = {'existant': 0, 'projet': 0, 'rdv': 0}
    
    # CORRECTION: Calcul des surfaces existantes (colonne "Surface existante avant travaux (A)")
    # Utiliser les polylignes de "Projet_demoli_feuille_TA.dxf" (dans main_existant_polylines)
    for polyline in main_existant_polylines:
        destination = get_destination_name(polyline)
        surface = calculer_surface_polyligne(polyline)
        surfaces_par_destination[destination]['existant'] += surface
    
    # CORRECTION: Calcul des surfaces projet (colonne "Surface projet")
    # Utiliser les polylignes de "Existant_exmple_demoli.dxf" (dans main_projet_polylines)
    for polyline in main_projet_polylines:
        destination = get_destination_name(polyline)
        surface = calculer_surface_polyligne(polyline)
        surfaces_par_destination[destination]['projet'] += surface
    
    # Méthode simplifiée pour déduire les surfaces des calques spéciaux
    logger.info("Début de la déduction des surfaces des calques spéciaux")
    
    # Pour les fichiers DXF, on sait que GEX_EDS_SDP_2, GEX_EDS_SDP_3, etc. sont des zones à déduire
    # des zones principales GEX_EDS_SDP_1 correspondantes
    
    # Structure pour garder trace des calques spéciaux traités
    special_layers_processed = {
        'existant': {},
        'projet': {}
    }
    
    # Traitement des polylignes spéciales dans le fichier existant
    try:
        logger.info(f"Traitement de {len(special_existant_polylines)} polylignes spéciales dans le fichier existant")
        
        for special_polyline in special_existant_polylines:
            special_layer = special_polyline.get('layer', '')
            if not isinstance(special_layer, str):
                continue
                
            logger.info(f"Traitement du calque spécial existant: {special_layer}")
            
            # Calculer la surface de cette polyligne spéciale
            special_surface = calculer_surface_polyligne(special_polyline)
            if special_surface <= 0:
                logger.warning(f"Surface nulle ou négative pour le calque {special_layer}, ignoré")
                continue
                
            # Déterminer le type de calque spécial (SDP_2, SDP_3, etc.)
            if 'GEX_EDS_SDP_2' in special_layer:
                category = 'TREMIE'
            elif 'GEX_EDS_SDP_3' in special_layer:
                category = 'H-180'
            elif 'GEX_EDS_SDP_5' in special_layer:
                category = 'PK'
            elif 'GEX_EDS_SDP_7' in special_layer:
                category = 'LT'
            else:
                logger.warning(f"Type de calque spécial inconnu: {special_layer}, ignoré")
                continue
                
            # Essayer de trouver la destination parente directement via les coordonnées
            # Cette méthode est plus simple et plus robuste
            parent_found = False
            
            for main_polyline in main_existant_polylines:
                if is_contained(special_polyline, main_polyline):
                    parent_destination = get_destination_name(main_polyline)
                    if parent_destination in surfaces_par_destination:
                        # Déduire la surface spéciale de la surface existante
                        surfaces_par_destination[parent_destination]['existant'] -= special_surface
                        logger.info(f"Surface {special_surface} déduite de {parent_destination} (existant) - Calque: {special_layer}")
                        special_layers_processed['existant'][special_layer] = {
                            'surface': special_surface,
                            'parent': parent_destination
                        }
                        parent_found = True
                        break
            
            if not parent_found:
                logger.warning(f"Aucun parent trouvé pour le calque {special_layer} dans l'existant")
                
    except Exception as e:
        logger.error(f"Erreur lors du traitement des calques spéciaux existants: {str(e)}")
    
    # Traitement des polylignes spéciales dans le fichier projet
    try:
        logger.info(f"Traitement de {len(special_projet_polylines)} polylignes spéciales dans le fichier projet")
        
        for special_polyline in special_projet_polylines:
            special_layer = special_polyline.get('layer', '')
            if not isinstance(special_layer, str):
                continue
                
            logger.info(f"Traitement du calque spécial projet: {special_layer}")
            
            # Calculer la surface de cette polyligne spéciale
            special_surface = calculer_surface_polyligne(special_polyline)
            if special_surface <= 0:
                logger.warning(f"Surface nulle ou négative pour le calque {special_layer}, ignoré")
                continue
                
            # Déterminer le type de calque spécial (SDP_2, SDP_3, etc.)
            if 'GEX_EDS_SDP_2' in special_layer:
                category = 'TREMIE'
            elif 'GEX_EDS_SDP_3' in special_layer:
                category = 'H-180'
            elif 'GEX_EDS_SDP_5' in special_layer:
                category = 'PK'
            elif 'GEX_EDS_SDP_7' in special_layer:
                category = 'LT'
            else:
                logger.warning(f"Type de calque spécial inconnu: {special_layer}, ignoré")
                continue
                
            # Essayer de trouver la destination parente directement via les coordonnées
            parent_found = False
            
            for main_polyline in main_projet_polylines:
                if is_contained(special_polyline, main_polyline):
                    parent_destination = get_destination_name(main_polyline)
                    if parent_destination in surfaces_par_destination:
                        # Déduire la surface spéciale de la surface projet
                        surfaces_par_destination[parent_destination]['projet'] -= special_surface
                        logger.info(f"Surface {special_surface} déduite de {parent_destination} (projet) - Calque: {special_layer}")
                        special_layers_processed['projet'][special_layer] = {
                            'surface': special_surface,
                            'parent': parent_destination
                        }
                        parent_found = True
                        break
            
            if not parent_found:
                logger.warning(f"Aucun parent trouvé pour le calque {special_layer} dans le projet")
                
    except Exception as e:
        logger.error(f"Erreur lors du traitement des calques spéciaux projet: {str(e)}")
    
    # Afficher un résumé des calques spéciaux traités
    logger.info(f"Résumé des calques spéciaux traités - Existant: {len(special_layers_processed['existant'])}, Projet: {len(special_layers_processed['projet'])}")

        
    # Importer les bibliothèques nécessaires pour les analyses spatiales
    try:
        from shapely.geometry import Polygon, Point
        shapely_available = True
    except ImportError:
        # Si Shapely n'est pas disponible, on affiche un message
        print("La bibliothèque Shapely n'est pas disponible. L'analyse spatiale précise ne sera pas possible.")
        shapely_available = False
        
    # Fonction pour convertir une polyligne en polygone Shapely
    def polyligne_to_polygon(polyline):
        if 'vertices' not in polyline or not polyline['vertices'] or len(polyline['vertices']) < 3:
            return None
        vertices = polyline['vertices']
        # Créer une liste de coordonnées (x, y) pour Shapely
        coords = [(vertex['x'], vertex['y']) for vertex in vertices]
        try:
            # Créer un polygone Shapely
            return Polygon(coords)
        except Exception as e:
            print(f"Erreur lors de la création du polygone: {e}")
            return None
    
    # Identifier les polylignes LOC_SOC et SANITAIRES dans tous les fichiers
    rdv_specific_polylines = []
    for polyline in projet_polylines + existant_polylines:
        layer = polyline.get('layer', '')
        if isinstance(layer, str) and ('LOC_SOC' in layer or 'SANITAIRES' in layer):
            rdv_specific_polylines.append(polyline)
    
    # Pour chaque destination (polyligne SDP), vérifier si elle contient des polylignes LOC_SOC ou SANITAIRES
    if shapely_available and rdv_specific_polylines:
        for destination in all_destinations:
            # Trouver toutes les polylignes associées à cette destination
            destination_polylines = []
            for polyline in main_projet_polylines + main_existant_polylines:
                if get_destination_name(polyline) == destination:
                    destination_polylines.append(polyline)
            
            # Pour chaque polyligne de destination, vérifier si elle contient des polylignes LOC_SOC ou SANITAIRES
            for dest_polyline in destination_polylines:
                dest_polygon = polyligne_to_polygon(dest_polyline)
                if not dest_polygon:
                    continue
                
                # Vérifier chaque polyligne spécifique
                for rdv_polyline in rdv_specific_polylines:
                    rdv_polygon = polyligne_to_polygon(rdv_polyline)
                    if not rdv_polygon:
                        continue
                    
                    # Vérifier si le polygone RDV est contenu dans le polygone de destination
                    if dest_polygon.contains(rdv_polygon) or dest_polygon.intersects(rdv_polygon):
                        # Calcul de la surface du polygone RDV
                        surface_rdv = rdv_polygon.area
                        # Ajouter cette surface à la destination
                        surfaces_par_destination[destination]['rdv'] += surface_rdv
        
    # Utiliser toutes les destinations collectées pour l'affichage
    destinations = all_destinations
    
    # Fonction pour formater les noms de destination pour un affichage plus lisible
    def format_destination(destination):
        # Dictionnaire de correspondance pour les cas spéciaux
        special_cases = {
            "AUTRE_BUREAU": "Autre bureau",
            "AUTRE_CONGRE": "Autre congrès exposition",
            "AUTRE_ENTREP": "Autre entrepôt",
            "AUTRE_INDUST": "Autre industrie",
            "COMMERCE_ART": "Commerce artisanat",
            "COMMERCE_AUT": "Commerce autre hébergement touristique",
            "COMMERCE_CIN": "Commerce cinéma",
            "COMMERCE_DE_": "Commerce de gros",
            "COMMERCE_HOT": "Commerce hôtel",
            "COMMERCE_RES": "Commerce restauration",
            "COMMERCE_SER": "Commerce service accueil clientèle",
            "EXPLOITATIO": "Exploitation forestière",
            "EXPLOITATION": "Exploitation agricole",
            "HABITATION_H": "Habitation hébergement",
            "HABITATION_L": "Habitation logement",
            "SPIC_ADMINIS": "Spic administration",
            "SPIC_ART_SPE": "Spic art spectacle",
            "SPIC_AUTRE": "Spic autre",
            "SPIC_ENSEIGN": "Spic enseignement santé",
            "SPIC_LT": "Spic lt",
            "SPIC_SPORT": "Spic sport"
        }
        
        # Si la destination est dans notre dictionnaire, utiliser la correspondance
        if destination in special_cases:
            return special_cases[destination]
        
        # Sinon, transformer le format générique: remplacer les underscores par des espaces et mettre la première lettre en majuscule
        formatted = destination.replace('_', ' ').lower().capitalize()
        return formatted
    
    # Trier les destinations par ordre alphabétique
    sorted_destinations = sorted(destinations)
    
    # Ajouter les destinations uniques à la feuille SDP avec leurs surfaces
    row = 2  # Commencer à la ligne 2
    
    # Stocker la première ligne et la dernière ligne pour fusionner les cellules plus tard
    first_row = row
    
    # Traiter chaque destination séparément
    for destination in sorted_destinations:
            
        # Colonne B: Destination (formatée pour une meilleure lisibilité)
        formatted_destination = format_destination(destination)
        ws_sdp[f'B{row}'] = formatted_destination
        
        # Appliquer le style à la cellule
        cell = ws_sdp[f'B{row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne C: Surface existante avant travaux (A)
        # Cette colonne est maintenant laissée intentionnellement vide selon la demande
        cell = ws_sdp[f'C{row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne I: Surface projet
        # Cette colonne est maintenant laissée intentionnellement vide selon la demande
        cell = ws_sdp[f'I{row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne J: Surface RDV
        surface_rdv = surfaces_par_destination.get(destination, {}).get('rdv', 0)
        if surface_rdv > 0:
            # Convertir en mètres carrés si nécessaire
            surface_rdv_m2 = surface_rdv
            
            # Arrondir à 2 décimales
            ws_sdp[f'J{row}'] = round(surface_rdv_m2, 2)
            
            # Appliquer le style à la cellule
            cell = ws_sdp[f'J{row}']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Ajouter un commentaire indiquant la source de la donnée
            cell.comment = Comment(f"Surface calculée à partir des polylignes RDV (GEX_EDS_RDV_)", "Calcul automatique")
            cell.comment.width = 300
            cell.comment.height = 50
        
        row += 1
    
    # Fusionner les cellules de la colonne A pour que le nom de l'étage n'apparaisse qu'une fois
    if row > 3:  # S'il y a plus d'une destination
        last_row = row - 1
        ws_sdp.merge_cells(f'A{first_row}:A{last_row}')
        
        # Centrer le nom de l'étage verticalement dans la cellule fusionnée
        cell = ws_sdp['A2']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Ajouter les destinations uniques à la feuille TA avec leurs surfaces
    ta_row = 2  # Commencer à la ligne 2
    
    # Dictionnaire pour stocker les données T.A. par destination
    ta_data = {}
    
    # Dictionnaire pour stocker les intersections détaillées des vides avec les zones
    intersection_details = {}
    
    # Identifier les polylignes pour le calcul du T.A.
    for destination in sorted_destinations:
        # Initialiser les valeurs pour cette destination
        planchers_avant_deductions = 0.0  # Planchers avant déductions
        vides = 0.0                       # Vides (stairwells, elevator shafts, etc.)
        surfaces_h_moins_180 = 0.0        # Surfaces dont h < 1.80m
        
        # Calculer Planchers avant déductions - utiliser les surfaces principales
        for polyline in main_projet_polylines:
            if get_destination_name(polyline) == destination:
                planchers_avant_deductions += calculer_surface_polyligne(polyline)
        
        # Identifier les vides (GEX_EDS_SDP_2 - VIDE)
        for polyline in special_projet_polylines:
            special_layer = polyline.get('layer', '')
            if not isinstance(special_layer, str):
                continue
            
            if 'GEX_EDS_SDP_2' in special_layer:  # VIDE
                vide_area = calculer_surface_polyligne(polyline)
                if vide_area <= 0:
                    continue
                
                # Identifiant unique pour ce vide
                vide_id = f"vide_{polyline.get('handle', '')}"
                if vide_id not in intersection_details:
                    intersection_details[vide_id] = {
                        'total_area': vide_area,
                        'intersections': {}
                    }
                    
                # CAS 1: Vérifier si ce vide est entièrement contenu dans cette zone
                contained_in_zone = False
                for main_polyline in main_projet_polylines:
                    if get_destination_name(main_polyline) == destination and is_contained(polyline, main_polyline):
                        vides += vide_area
                        intersection_details[vide_id]['intersections'][destination] = vide_area
                        logger.info(f"Excel: Vide {special_layer} entièrement contenu dans la zone {destination}, surface déduite: {vide_area:.2f} m²")
                        contained_in_zone = True
                        break
                        
                # CAS 2: Si le vide n'est pas entièrement contenu, vérifier s'il intersecte cette zone
                if not contained_in_zone:
                    # Détection spéciale pour congre_art
                    is_congre_art = False
                    if destination == "congre_art" or "congre_art" in destination.lower():
                        logger.warning(f"Détection de zone congre_art: {destination}")
                        is_congre_art = True
                    
                    for main_polyline in main_projet_polylines:
                        dest_name = get_destination_name(main_polyline)
                        if dest_name == destination:
                            # Débogage avancé pour les intersections
                            does_intersect = intersects(polyline, main_polyline)
                            if is_congre_art:
                                logger.warning(f"Vérification d'intersection avec congre_art: {does_intersect}")
                                # Force l'intersection pour congre_art si nécessaire
                                if not does_intersect:
                                    # Vérifier si les boîtes englobantes se chevauchent
                                    from shapely.geometry import box
                                    
                                    # Créer les boîtes englobantes
                                    poly1_pts = [(v['x'], v['y']) for v in polyline.get('vertices', [])]
                                    poly2_pts = [(v['x'], v['y']) for v in main_polyline.get('vertices', [])]
                                    
                                    if len(poly1_pts) >= 3 and len(poly2_pts) >= 3:
                                        min_x1, min_y1 = min(p[0] for p in poly1_pts), min(p[1] for p in poly1_pts)
                                        max_x1, max_y1 = max(p[0] for p in poly1_pts), max(p[1] for p in poly1_pts)
                                        min_x2, min_y2 = min(p[0] for p in poly2_pts), min(p[1] for p in poly2_pts)
                                        max_x2, max_y2 = max(p[0] for p in poly2_pts), max(p[1] for p in poly2_pts)
                                        
                                        box1 = box(min_x1, min_y1, max_x1, max_y1)
                                        box2 = box(min_x2, min_y2, max_x2, max_y2)
                                        
                                        if box1.intersects(box2):
                                            logger.warning(f"Les boîtes englobantes se chevauchent mais l'intersection géométrique a échoué - forçage de l'intersection pour congre_art")
                                            does_intersect = True
                            
                            if dest_name == destination and does_intersect:
                                # Calculer l'intersection exacte entre le vide et cette zone
                                intersection_area = calculate_intersection(polyline, main_polyline)
                                if intersection_area > 0:
                                    # Vérifier que l'intersection n'est pas déjà comptabilisée pour une autre zone
                                    # (cas où le vide croise plusieurs zones)
                                    already_counted = sum(intersection_details[vide_id]['intersections'].get(dest, 0) 
                                                         for dest in intersection_details[vide_id]['intersections'])
                                    
                                    # S'assurer que la somme des intersections ne dépasse pas la surface totale du vide
                                    if already_counted + intersection_area > vide_area * 1.01:  # 1% de tolérance pour les erreurs de calcul
                                        adjusted_area = max(0, vide_area - already_counted)
                                        logger.warning(f"Ajustement de l'intersection pour éviter le double comptage: {intersection_area:.2f} → {adjusted_area:.2f} m²")
                                        intersection_area = adjusted_area
                                    
                                    vides += intersection_area
                                    intersection_details[vide_id]['intersections'][destination] = intersection_area
                                    logger.info(f"Excel: Vide {special_layer} intersecte la zone {destination}, surface déduite: {intersection_area:.2f} m² (sur {vide_area:.2f} m² total)")
                                    
                                    # Vérifier que la déduction n'excède pas la surface disponible
                                    if intersection_area > planchers_avant_deductions:
                                        logger.warning(f"ATTENTION: Déduction de vide ({intersection_area:.2f} m²) supérieure à la surface disponible ({planchers_avant_deductions:.2f} m²) pour {destination}")
                                        intersection_area = min(intersection_area, planchers_avant_deductions)
                                        vides = planchers_avant_deductions - (surfaces_h_moins_180)  # Limiter la déduction à la surface disponible

        
        # Identifier les surfaces dont h < 1.80m (GEX_EDS_SDP_3 - H-180)
        for polyline in special_projet_polylines:
            special_layer = polyline.get('layer', '')
            if not isinstance(special_layer, str):
                continue
            
            if 'GEX_EDS_SDP_3' in special_layer:  # H-180
                h180_area = calculer_surface_polyligne(polyline)
                if h180_area <= 0:
                    continue
                
                # Identifiant unique pour cette surface H-180
                h180_id = f"h180_{polyline.get('handle', '')}"
                if h180_id not in intersection_details:
                    intersection_details[h180_id] = {
                        'total_area': h180_area,
                        'intersections': {}
                    }
                    
                # CAS 1: Vérifier si cette surface H-180 est entièrement contenue dans cette zone
                contained_in_zone = False
                for main_polyline in main_projet_polylines:
                    if get_destination_name(main_polyline) == destination and is_contained(polyline, main_polyline):
                        surfaces_h_moins_180 += h180_area
                        intersection_details[h180_id]['intersections'][destination] = h180_area
                        logger.info(f"Excel: Surface H-180 {special_layer} entièrement contenue dans la zone {destination}, surface déduite: {h180_area:.2f} m²")
                        contained_in_zone = True
                        break
                        
                # CAS 2: Si la surface H-180 n'est pas entièrement contenue, vérifier si elle intersecte cette zone
                if not contained_in_zone:
                    for main_polyline in main_projet_polylines:
                        if get_destination_name(main_polyline) == destination and intersects(polyline, main_polyline):
                            # Calculer l'intersection exacte entre la surface H-180 et cette zone
                            intersection_area = calculate_intersection(polyline, main_polyline)
                            if intersection_area > 0:
                                # Vérifier que l'intersection n'est pas déjà comptabilisée pour une autre zone
                                # (cas où la surface H-180 croise plusieurs zones)
                                already_counted = sum(intersection_details[h180_id]['intersections'].get(dest, 0) 
                                                     for dest in intersection_details[h180_id]['intersections'])
                                
                                # S'assurer que la somme des intersections ne dépasse pas la surface totale de H-180
                                if already_counted + intersection_area > h180_area * 1.01:  # 1% de tolérance pour les erreurs de calcul
                                    adjusted_area = max(0, h180_area - already_counted)
                                    logger.warning(f"Ajustement de l'intersection H-180 pour éviter le double comptage: {intersection_area:.2f} → {adjusted_area:.2f} m²")
                                    intersection_area = adjusted_area
                                
                                surfaces_h_moins_180 += intersection_area
                                intersection_details[h180_id]['intersections'][destination] = intersection_area
                                logger.info(f"Excel: Surface H-180 {special_layer} intersecte la zone {destination}, surface déduite: {intersection_area:.2f} m² (sur {h180_area:.2f} m² total)")
                                
                                # Vérifier que la déduction n'excède pas la surface disponible
                                if intersection_area > planchers_avant_deductions:
                                    logger.warning(f"ATTENTION: Déduction de surface H-180 ({intersection_area:.2f} m²) supérieure à la surface disponible ({planchers_avant_deductions:.2f} m²) pour {destination}")
                                    intersection_area = min(intersection_area, planchers_avant_deductions)
                                    surfaces_h_moins_180 = min(surfaces_h_moins_180, planchers_avant_deductions - vides)  # Limiter la déduction à la surface disponible
        
        # Recalculer les vides et surfaces h-180 à partir des intersections détaillées
        # pour s'assurer que les ajustements sont bien pris en compte
        vides_from_intersections = 0
        for vide_id, details in intersection_details.items():
            if vide_id.startswith('vide_') and destination in details['intersections']:
                vides_from_intersections += details['intersections'][destination]
        
        h180_from_intersections = 0
        for h180_id, details in intersection_details.items():
            if h180_id.startswith('h180_') and destination in details['intersections']:
                h180_from_intersections += details['intersections'][destination]
        
        # Utiliser les valeurs calculées à partir des intersections
        vides = vides_from_intersections
        surfaces_h_moins_180 = h180_from_intersections
        
        # Vérifier que les déductions ne dépassent pas la surface disponible
        if vides + surfaces_h_moins_180 > planchers_avant_deductions * 1.01:  # 1% de tolérance
            logger.warning(f"ATTENTION: Déductions totales ({vides + surfaces_h_moins_180:.2f} m²) supérieures à la surface disponible ({planchers_avant_deductions:.2f} m²) pour {destination}")
            # Ajuster proportionnellement
            if vides + surfaces_h_moins_180 > 0:
                ratio = planchers_avant_deductions / (vides + surfaces_h_moins_180)
                vides *= ratio
                surfaces_h_moins_180 *= ratio
                logger.warning(f"Ajustement proportionnel des déductions: vides={vides:.2f} m², h-180={surfaces_h_moins_180:.2f} m²")
        
        # Calculer le TOTAL T.A.
        total_ta = planchers_avant_deductions - (vides + surfaces_h_moins_180)
        
        # S'assurer que le total n'est pas négatif
        total_ta = max(0, total_ta)
        
        # Stocker les données pour cette destination
        ta_data[destination] = {
            'planchers_avant_deductions': planchers_avant_deductions,
            'vides': vides,
            'surfaces_h_moins_180': surfaces_h_moins_180,
            'total_ta': total_ta
        }
    
    # Ajouter les données T.A. à la feuille TA
    ta_row = 2  # Commencer à la ligne 2
    for destination in sorted_destinations:
        # Colonne B: Destination (formatée pour une meilleure lisibilité)
        formatted_destination = format_destination(destination)
        ws_ta[f'B{ta_row}'] = formatted_destination
        
        # Appliquer le style à la cellule
        cell = ws_ta[f'B{ta_row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Récupérer les données T.A. pour cette destination
        destination_ta_data = ta_data.get(destination, {})
        
        # Colonne C: TA avant déduction (renommé de "Planchers avant déductions")
        planchers_avant_deductions = destination_ta_data.get('planchers_avant_deductions', 0)
        ws_ta[f'C{ta_row}'] = round(planchers_avant_deductions, 4) if planchers_avant_deductions > 0 else 0
        cell = ws_ta[f'C{ta_row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne D: Vides
        vides = destination_ta_data.get('vides', 0)
        ws_ta[f'D{ta_row}'] = round(vides, 4) if vides > 0 else 0
        cell = ws_ta[f'D{ta_row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Ajouter un commentaire détaillant les intersections des vides avec cette zone
        vide_details = []
        for vide_id, details in intersection_details.items():
            if vide_id.startswith('vide_') and destination in details['intersections']:
                intersection_area = details['intersections'][destination]
                total_area = details['total_area']
                vide_details.append(f"Vide {vide_id.replace('vide_', '')}: {round(intersection_area, 2)} m² (sur {round(total_area, 2)} m² total)")
        
        if vide_details:
            comment_text = "Détail des intersections de vides:\n" + "\n".join(vide_details)
            comment = Comment(comment_text, "Détail des vides")
            comment.width = 400
            comment.height = 100 + (len(vide_details) * 15)  # Ajuster la hauteur en fonction du nombre de lignes
            cell.comment = comment
        
        # Colonne E: Surfaces dont h < 1.80m
        surfaces_h_moins_180 = destination_ta_data.get('surfaces_h_moins_180', 0)
        ws_ta[f'E{ta_row}'] = round(surfaces_h_moins_180, 4) if surfaces_h_moins_180 > 0 else 0
        cell = ws_ta[f'E{ta_row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Ajouter un commentaire détaillant les intersections des surfaces H-180 avec cette zone
        h180_details = []
        for h180_id, details in intersection_details.items():
            if h180_id.startswith('h180_') and destination in details['intersections']:
                intersection_area = details['intersections'][destination]
                total_area = details['total_area']
                h180_details.append(f"Surface H-180 {h180_id.replace('h180_', '')}: {round(intersection_area, 2)} m² (sur {round(total_area, 2)} m² total)")
        
        if h180_details:
            comment_text = "Détail des intersections de surfaces H-180:\n" + "\n".join(h180_details)
            comment = Comment(comment_text, "Détail des surfaces H-180")
            comment.width = 400
            comment.height = 100 + (len(h180_details) * 15)  # Ajuster la hauteur en fonction du nombre de lignes
            cell.comment = comment
        
        # Colonne F: TA après déduction (renommé de "TOTAL T.A.")
        total_ta = destination_ta_data.get('total_ta', 0)
        ws_ta[f'F{ta_row}'] = round(total_ta, 4) if total_ta > 0 else 0
        cell = ws_ta[f'F{ta_row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Ajouter un commentaire expliquant le calcul
        cell.comment = Comment(f"TA après déduction = TA avant déduction - (Vides + Surfaces dont h < 1.80m)", "Calcul T.A.")
        cell.comment.width = 300
        cell.comment.height = 50
        
        # Colonne G: Pourcentage de vide par rapport à la surface avant déductions
        vide_percentage = 0
        if planchers_avant_deductions > 0:
            vide_percentage = (vides / planchers_avant_deductions) * 100
        
        ws_ta[f'G{ta_row}'] = f"{vide_percentage:.2f}%"
        cell = ws_ta[f'G{ta_row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Ajouter un avertissement visuel si la déduction est trop importante
        if vide_percentage > 50:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.comment = Comment(f"ATTENTION: Déduction de vide supérieure à 50% de la surface totale", "Alerte")
            cell.comment.width = 200
            cell.comment.height = 50
        
        ta_row += 1
    
    # Création de la feuille TA Summary avec informations d'intersection de vides
    ws_ta_summary = wb.create_sheet(title="TA Summary - VIDES INTERSECTIONS")
    
    # Configuration des colonnes TA Summary
    for col in columns:
        ws_ta_summary.column_dimensions[col].width = 15
    
    # En-tête de la feuille TA Summary
    ws_ta_summary['A1'] = "Étages"
    ws_ta_summary['B1'] = "Total TA avant déduction"
    ws_ta_summary['C1'] = "Total Vides"
    ws_ta_summary['D1'] = "Total H-180"
    ws_ta_summary['E1'] = "Total TA après déduction"
    ws_ta_summary['F1'] = "% Vides"
    ws_ta_summary['G1'] = "TA Projet"
    ws_ta_summary['H1'] = "TA Existant"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws_ta_summary[f'{col}1']
        cell.font = subheader_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Calculer les totaux pour le résumé
    total_ta_avant_deduction = 0
    total_vides = 0
    total_h_moins_180 = 0
    total_ta_apres_deduction = 0
    
    # Calculer les totaux pour TA Projet et TA Existant
    ta_projet_total = 0
    ta_existant_total = 0
    
    # Parcourir les polylignes pour calculer les totaux
    projet_polylines = surfaces.get('projet', {}).get('polylines', [])
    existant_polylines = surfaces.get('existant', {}).get('polylines', [])
    
    for polyline in projet_polylines:
        if is_main_sdp_polyline(polyline):
            ta_projet_total += calculer_surface_polyligne(polyline)
            
    for polyline in existant_polylines:
        if is_main_sdp_polyline(polyline):
            ta_existant_total += calculer_surface_polyligne(polyline)
    
    # Parcourir toutes les destinations pour calculer les totaux
    for destination in sorted_destinations:
        destination_ta_data = ta_data.get(destination, {})
        total_ta_avant_deduction += destination_ta_data.get('planchers_avant_deductions', 0)
        total_vides += destination_ta_data.get('vides', 0)
        total_h_moins_180 += destination_ta_data.get('surfaces_h_moins_180', 0)
        total_ta_apres_deduction += destination_ta_data.get('total_ta', 0)
    
    # Calculer le pourcentage total de vides
    total_vide_percentage = 0
    if total_ta_avant_deduction > 0:
        total_vide_percentage = (total_vides / total_ta_avant_deduction) * 100
    
    # Ajouter les données à la feuille TA Summary
    ws_ta_summary['A2'] = floor_name
    ws_ta_summary['B2'] = round(total_ta_avant_deduction, 4) if total_ta_avant_deduction > 0 else 0
    ws_ta_summary['C2'] = round(total_vides, 4) if total_vides > 0 else 0
    ws_ta_summary['D2'] = round(total_h_moins_180, 4) if total_h_moins_180 > 0 else 0
    ws_ta_summary['E2'] = round(total_ta_apres_deduction, 4) if total_ta_apres_deduction > 0 else 0
    ws_ta_summary['F2'] = f"{total_vide_percentage:.2f}%"
    ws_ta_summary['G2'] = round(ta_projet_total, 4) if ta_projet_total > 0 else 0
    ws_ta_summary['H2'] = round(ta_existant_total, 4) if ta_existant_total > 0 else 0
    
    # Appliquer le style aux cellules
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws_ta_summary[f'{col}2']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
    
    # Sauvegarde du fichier Excel
    wb.save(excel_path)


